from datetime import datetime
from openpyxl import Workbook
# from dxgui import dxdialogs
import dxbuild.buildtools as buildtools
from dxbuild.buildtools import autoincrement_name

# Debug information
_WRITE_FILE = False
_SHEET_BASE_NAME = 'Comments'

# Create Workbook and add auto-incremented sheet names
wb = Workbook()

def generate_sheets(wb: Workbook, names_list: list[str]):
    if len(wb.sheetnames) == 1:
        ws = wb.active
        ws.title = names_list[0]
    for i in range(1, len(names_list)):
        new_ws_name = autoincrement_name(names_list[i], wb.sheetnames)
        wb.create_sheet(new_ws_name)

names_list = ['Comments', 'Comments', 'Comments', 'Comments']
alt_names_list = ['Concept', 'Intermediate', 'Final', 'Backcheck']

generate_sheets(wb, alt_names_list)
print(wb.sheetnames)

if _WRITE_FILE:
    save_name = f'./dev/test/out/test_{buildtools.timestamp()}.xlsx'
    wb.save(save_name)
    print(f'{save_name} has been saved.')
wb.close()


