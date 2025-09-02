from drchecks_reviews import Review
from utils import timestamp, list_dimensions
import pandas as pd
import numpy as np
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

xml_path = 'test.xml'
review = Review.from_file(xml_path)
project_info = review.project_info
review_comments = review.review_comments

expansion_type = 'type'
all_list, rows = review_comments.get_all_comments_and_responses(expansion_type=expansion_type)
header_list, columns = review_comments.get_all_comments_and_response_headers(expansion_type=expansion_type)

# df = pd.DataFrame(all_list, columns=header_list)
# df.to_excel(f'all_output_{expansion_type}_{timestamp()}.xlsx', index=False)


# df2 = pd.DataFrame(project_info.get_info)
# print(df2)

# a_comment = review_comments.comments[1]
# print(a_comment.latest_response.remark_type)
# print(a_comment.ball_in_court)

# for comment in review_comments.comments:
#     print(comment.id, comment.ball_in_court)

# print(rows, columns)

# np_list = np.array(header_list)
# print(np_list.shape)

def transpose_data(data_list: List) -> List:
    rows, columns = list_dimensions(data_list)
    return [[[j][i] for j in range(rows)] for i in range(columns)]

def get_list_range(data_list: List, 
                   anchor_cell: str='A1', 
                   to_transposed: bool=False, 
                   return_R1C1: bool=False) -> str | Tuple[int, int, int, int]:
    # TODO: This function is broken...
    anchor_row, anchor_column = coordinate_to_tuple(anchor_cell)
    rows, columns = list_dimensions(data_list)
    # columns -= 1
    if to_transposed:
        rows, columns = columns, rows
    if return_R1C1:
        return (anchor_row, anchor_column, anchor_row + rows, anchor_column + columns)
    # Calculate the bottom-right cell
    end_row = anchor_row + rows
    end_col = anchor_column + columns
    end_cell = f"{get_column_letter(end_col)}{end_row}"
    return f"{anchor_cell}:{end_cell}"


def copy_to_range(data_list: List, 
                  worksheet: Worksheet | None, 
                  anchor_cell: str='A1'):
    if worksheet is not None:
        anchor_row, anchor_column = coordinate_to_tuple(anchor_cell)
        rows, columns = list_dimensions(data_list)
        if rows == 1:
            for j in range(columns):
                worksheet.cell(row=anchor_row, column=anchor_column + j).value = data_list[j]
        else:
            for i in range(rows):
                for j in range(columns):
                    worksheet.cell(row=anchor_row + i, column=anchor_column + j).value = data_list[i][j]



xlsx_path = f'export_test_{timestamp()}.xlsx'

wb = Workbook()
ws = wb.active


HEADER_CELL = 'H11'
DATABODY_CELL = 'I11'

print(get_list_range(all_list, 'I11', to_transposed=True, return_R1C1=False))
print(get_list_range(all_list, 'I11', to_transposed=True, return_R1C1=True))

anchor_row, anchor_column = coordinate_to_tuple(HEADER_CELL)
rows, columns = list_dimensions(all_list)

copy_to_range(transpose_data(header_list), worksheet=ws, anchor_cell=HEADER_CELL)
# copy_to_range(all_list, worksheet=ws, anchor_cell='I11', is_transposed=True)

# table_range = CellRange(min_row=anchor_row,
#                         min_col=anchor_column,
#                         max_row=anchor_row+columns - 1,
#                         max_col=anchor_column+rows)

# print(table_range.coord)

# table_style = TableStyleInfo(name="TableStyleMedium9", 
#                              showRowStripes=False, 
#                              showFirstColumn=False, 
#                              showLastColumn=False,
#                              showColumnStripes=False)

# table = Table(displayName='Comments',
#               ref=table_range.coord)

# table.tableStyleInfo = table_style
# if ws is not None:
#     ws.add_table(table)

wb.save(xlsx_path)