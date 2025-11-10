# Copyright (c) 2018-2025 Ben Fisher

import os, sys, shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import DEFAULT_FONT, Alignment
from openpyxl.utils.cell import get_column_letter

from PyQt6.QtWidgets import QFileDialog
from .bidconstants import FALLBACKS, timestamp, BID_TABLE_HEADERS
from .bidhtml import read_bid_html_to_list

import dxbuild.buildtools as buildtools
from dxcore.cellformats import BID_LOG_COLUMN_WIDTHS, table_body_wrap_styles, table_body_styles

_template_path = './bid/template.xlsx'

def create_bid_log() -> str | bool:

    _WRITE_FILE = True
    DEFAULT_FONT.__init__(name=FALLBACKS['font_name'], size=FALLBACKS['font_size'])
    html_path, _ = QFileDialog.getOpenFileName(
        parent=None, 
        caption='Select File Dialog',
        filter='HTML Files (*.html);;All Files (*)'
    )

    copied_file_path = os.path.join(os.path.dirname(html_path), f'Bidder RFI Log {timestamp('%Y-%m-%d %H-%M-%S')}.xlsx')
    copied_file = shutil.copy(_template_path, copied_file_path)

    try:
        wb = load_workbook(copied_file)
    except FileNotFoundError as e:
        print(f'Could not find the template file: {e}')
        exit()

    # Check to see if the user selected any files or cancelled the file select dialog.
    if len(html_path) > 0:
        bidder_rfis = read_bid_html_to_list(html_path=html_path)
        if bidder_rfis:

            print('Will create new sheet')
            ws = wb.create_sheet('Bidder RFIs', index=3)
            ws.sheet_view.showGridLines = False

            comments_list = read_bid_html_to_list(html_path)

            buildtools.copy_to_range(BID_TABLE_HEADERS, ws, 'A9')
            buildtools.copy_to_range(comments_list, ws, 'A10')

            for i, col_width in enumerate(BID_LOG_COLUMN_WIDTHS):
                ws.column_dimensions[get_column_letter(i + 1)].width = col_width

            top_left_alignment = Alignment(horizontal='left', vertical='top')
            used_range = ws.calculate_dimension()
            for row in ws[used_range]:
                for cell in row:
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'm/d/yy'
                    cell.format = table_body_wrap_styles

        else:
            print('No bidder RFIs found in the HTML file.')
            _WRITE_FILE = False

        if _WRITE_FILE:
            wb.save(copied_file_path)
        wb.close()
        print(f'{copied_file_path} written to disk.')
        os.startfile(os.path.dirname(copied_file_path))
        # os.startfile(copied_file_path)
        return copied_file_path
    else:
        return False
        