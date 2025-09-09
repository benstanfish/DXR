from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, DEFAULT_FONT
import openpyxl.utils

file_path = './output.xlsx'

# normal_font = Font(name='Arial', size=14)
# openpyxl.styles.DEFAULT_FONT.font = Font(name='Arial', size=14)
# DEFAULT_FONT.__init__(name='Aptos Narrow', size=15)

# openpyxl.styles.DEFAULT_FONT.name = 'Aptos'
# print(DEFAULT_FONT.name)


wb = Workbook()

# wb._fonts[0].name = 'Arial'

ws = wb.active
ws.title = 'sheet_title'

tree_data = [["Type", "Leaf Color", "Height"],
            ["Maple", "Red", 549],
            ["Oak", "Green", 783],
            ["Pine", "Green", 1204]]

for row in tree_data:
    ws.append(row)

# normal_font = Font(name='Aptos',
#                    size=11)

# header_font = Font(name='Aptos',
#                    size=11,
#                    bold=True)

# regular_align = Alignment(horizontal='left',
#                           vertical='top')

# all_cells = ws['A1:C4']
# for row in all_cells:
#     for cell in row:
#         cell.font = normal_font
#         cell.alignment = regular_align

# header_range = ws['A1:C1']
# for row in header_range:
#     for cell in row:
#         cell.font = header_font
#         cell.alignment = regular_align




wb.save(file_path)
wb.close()