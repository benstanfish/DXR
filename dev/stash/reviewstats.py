# Copyright (c) 2018-2025 Ben Fisher

import sys, os
# import win32com.client as COM

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import DEFAULT_FONT, Font, Alignment
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter

from PyQt6.QtWidgets import QApplication, QFileDialog

from dxbuild.reviews import Review
from dxbuild.constants import FALLBACKS, _LOG_DIR
from dxbuild.buildtools import timestamp, copy_to_range, apply_styles_to_region
from dxcore.cellformats import *
from dxreport import singlereport

if not os.path.exists(_LOG_DIR):
    os.makedirs(_LOG_DIR)

import logging
from dxcore.logconstants import log_format_string
logger = logging.getLogger(__name__)
logger.setLevel(logging.WARNING)
log_formatter = logging.Formatter(log_format_string)
log_file_handler = logging.FileHandler(f'{_LOG_DIR}/{__name__}.log')
log_file_handler.setFormatter(log_formatter)
logger.addHandler(log_file_handler)

_PADDING_OFFSET = 1

# def main() -> None:

    # _WRITE_FILE = True
    # xml_path = './dev/test/data.xml'

    # DEFAULT_FONT.__init__(name=FALLBACKS['font_name'], size=FALLBACKS['font_size'])
    # wb = Workbook()


    # review = Review.from_file(xml_path)
    

    # ws = wb.active
    
    # singlereport.create_report(review, ws)

def make_stats_sheet(ws: Worksheet, review: Review) -> None:

    wb = ws.parent

    ws_table_name = ws.tables.items()[0][0]
    ws_stat_name = ws.title[:25] + '-STAT'

    ws_stats = wb.create_sheet(ws_stat_name)
    ws_stats.sheet_view.showGridLines = False
    
    comments = review.review_comments.comments

    disciplines = list(set([comment.discipline for comment in comments]))
    disciplines.sort()

    reviewers = list(set([comment.author for comment in comments]))
    reviewers.sort()

    highest_responses = list(set([comment.highest_response() for comment in comments]))
    highest_responses.sort()

    reviewers_with_open_comments = list(set([comment.author for comment in comments if comment.status == 'Open']))
    reviewers_with_open_comments.sort()
    
    reviewer_open_comments_dict = {}
    for reviewer in reviewers_with_open_comments:
        comment_ids = list(set([comment.id for comment in comments if (comment.status == 'Open' and comment.author == reviewer)]))
        reviewer_open_comments_dict[reviewer] = comment_ids


    project_header_anchor = CellRange('A1')
    project_header = [
        ['Dr Checks Review Statistics',''],
        ['', ''],
        ['Project Name', review.project_info.project_name],
        ['Project ID', review.project_info.project_id],
        ['Review Name', review.project_info.review_name],
    ]
    copy_to_range(project_header, ws_stats, project_header_anchor.coord)
    

    overall_status_anchor = ws_stats[project_header_anchor.coord].offset(row=len(project_header) + _PADDING_OFFSET, column=0)
    overall_status = [
        ['Overall Comment Status', '', '', ''],
        ['By Discipline', 'Open', 'Closed', 'Total']
    ]
    start_row = CellRange(overall_status_anchor.coordinate).min_row + len(overall_status)
    start_column = CellRange(overall_status_anchor.coordinate).min_col
    for i, discipline in enumerate(disciplines):
        temp_row = []
        temp_row.append(discipline)
        temp_row.append(f'=COUNTIFS({ws_table_name}[Discipline],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 1)}${start_row - 1})')
        temp_row.append(f'=COUNTIFS({ws_table_name}[Discipline],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 2)}${start_row - 1})')
        temp_row.append(f'=SUM({get_column_letter(start_column + 1)}{start_row + i}:{get_column_letter(start_column + 2)}{start_row + i})')
        overall_status.append(temp_row)
    temp_totals_row = ['Grand Total']
    for i in range(1, 4):
        temp_totals_row.append(f'=SUM({get_column_letter(start_column + i)}{start_row}:{get_column_letter(start_column + i)}{start_row + len(disciplines) - 1})')
    overall_status.append(temp_totals_row)
    copy_to_range(overall_status, ws_stats, overall_status_anchor.coordinate)

    
    reviewer_status_anchor = ws_stats[project_header_anchor.coord].offset(row=len(project_header) + _PADDING_OFFSET + 1, column=4 + _PADDING_OFFSET)
    reviewer_status = [
        ['By Author', 'Open', 'Closed', 'Total']
    ]
    start_row = CellRange(reviewer_status_anchor.coordinate).min_row + len(reviewer_status)
    start_column = CellRange(reviewer_status_anchor.coordinate).min_col    
    for i, reviewer in enumerate(reviewers):
        temp_row = []
        temp_row.append(reviewer)
        temp_row.append(f'=COUNTIFS({ws_table_name}[Author],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 1)}${start_row - 1})')
        temp_row.append(f'=COUNTIFS({ws_table_name}[Author],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 2)}${start_row - 1})')
        temp_row.append(f'=SUM({get_column_letter(start_column + 1)}{start_row + i}:{get_column_letter(start_column + 2)}{start_row + i})')
        reviewer_status.append(temp_row)
    temp_totals_row = ['Grand Total']
    for i in range(1, 4):
        temp_totals_row.append(f'=SUM({get_column_letter(start_column + i)}{start_row}:{get_column_letter(start_column + i)}{start_row + len(reviewers) - 1})')
    reviewer_status.append(temp_totals_row)
    copy_to_range(reviewer_status, ws_stats, reviewer_status_anchor.coordinate)    


    try:
        response_status_anchor = reviewer_status_anchor.offset(row=0, column=4 + _PADDING_OFFSET)
        response_status = [
            ['By Response', 'Open', 'Closed', 'Total']
        ]
        start_row = CellRange(response_status_anchor.coordinate).min_row + len(response_status)
        start_column = CellRange(response_status_anchor.coordinate).min_col  
        for i, highest_response in enumerate(highest_responses):
            temp_row = []
            if highest_response == '':
                temp_row.append('None')
                temp_row.append(f'=COUNTIFS({ws_table_name}[Highest Resp.],"",{ws_table_name}[Status],{get_column_letter(start_column + 1)}${start_row - 1})')
                temp_row.append(f'=COUNTIFS({ws_table_name}[Highest Resp.],"",{ws_table_name}[Status],{get_column_letter(start_column + 2)}${start_row - 1})')
            else:
                temp_row.append(highest_response)
                temp_row.append(f'=COUNTIFS({ws_table_name}[Highest Resp.],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 1)}${start_row - 1})')
                temp_row.append(f'=COUNTIFS({ws_table_name}[Highest Resp.],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 2)}${start_row - 1})')
            temp_row.append(f'=SUM({get_column_letter(start_column + 1)}{start_row + i}:{get_column_letter(start_column + 2)}{start_row + i})')
            response_status.append(temp_row)
        temp_totals_row = ['Grand Total']
        for i in range(1, 4):
            temp_totals_row.append(f'=SUM({get_column_letter(start_column + i)}{start_row}:{get_column_letter(start_column + i)}{start_row + len(highest_responses) - 1})')
        response_status.append(temp_totals_row)
        copy_to_range(response_status, ws_stats, response_status_anchor.coordinate)   
    except Exception as e:
        print(f"Couldn't write the response region; likely no responses yet. Error at {e}.")
        logger.exception(f"Couldn't write the response region; likely no responses yet. Error at {e}.")


    status_region_max_rows = max(len(overall_status), len(reviewer_status))

    open_comments_anchor = ws_stats[project_header_anchor.coord].offset(row=len(project_header) + status_region_max_rows + _PADDING_OFFSET * 2 + 1, column=0)
    open_comments_anchor.value = 'Open Comments by Author'
    for i, reviewer in enumerate(reviewer_open_comments_dict.keys()):
        open_comments_anchor.offset(row=1, column=i).value = reviewer
        comment_id_list = reviewer_open_comments_dict[reviewer]
        comment_id_list.sort()
        for j, comment_id in enumerate(comment_id_list):
            open_comments_anchor.offset(row=2 + j, column=i).value = comment_id



    # Global formatting: reset
    for row in ws_stats.iter_rows():
        for cell in row:
            cell.font = Font(name='Aptos Narrow', sz=10)
            cell.alignment = Alignment(horizontal='left')

    # Lower priority column width changes first
    for i, _ in enumerate(reviewer_open_comments_dict.keys()):
        ws_stats.column_dimensions[get_column_letter(open_comments_anchor.offset(column=i).column)].width = Widths.SMALL

    # Format the title and region titles
    ws_stats[project_header_anchor.coord].font = Font(name='Aptos', sz=14, bold=True)
    ws_stats[overall_status_anchor.coordinate].font = Font(name='Aptos', sz=12, bold=True)
    ws_stats[open_comments_anchor.coordinate].font = Font(name='Aptos', sz=12, bold=True)

    ws_stats.column_dimensions[get_column_letter(overall_status_anchor.column)].width = Widths.SMALL
    ws_stats.column_dimensions[get_column_letter(reviewer_status_anchor.column)].width = Widths.SMALL
    try:
        ws_stats.column_dimensions[get_column_letter(response_status_anchor.column)].width = Widths.SMALL
    except Exception as e:
        logger(f"Couldn't change the width of the Response Status region; perhaps there isn't a Response region? Error: {e}")


    # Format the bold fields in the project header region
    for i in range(2, len(project_header)):
        ws_stats[project_header_anchor.coord].offset(row=i, column=0).font = Font(name='Aptos Narrow', sz=10, bold=True)
        if i == 2:
            ws_stats[project_header_anchor.coord].offset(row=i, column=1).font = Font(name='Aptos Narrow', sz=12, bold=True)


    # Format the headers of each overall regions
    def create_formatting_regions(initial_anchor: str, region_width: int) -> str:
        temp_range = CellRange(range_string=initial_anchor)
        temp_range.expand(right=region_width)
        return temp_range.coord

    headers = []
    headers.append(create_formatting_regions(overall_status_anchor.offset(row=1).coordinate, len(overall_status[0]) - 1))
    headers.append(create_formatting_regions(reviewer_status_anchor.coordinate, len(reviewer_status[0]) - 1))
    try:
        headers.append(create_formatting_regions(response_status_anchor.coordinate, len(response_status[0]) - 1))
    except Exception as e:
        logger(f"Couldn't format the Response Status region header; perhaps there isn't a Response region? Error: {e}") 
    headers.append(create_formatting_regions(open_comments_anchor.offset(row=1).coordinate, len(reviewer_open_comments_dict.keys()) - 1))
    for range_string in headers:
        apply_styles_to_region(stat_header_region_styles, range_string, ws_stats)


    footers = []
    footers.append(create_formatting_regions(overall_status_anchor.offset(row=len(overall_status) - 1).coordinate, len(overall_status[0]) - 1))
    footers.append(create_formatting_regions(reviewer_status_anchor.offset(row=len(reviewer_status) - 1).coordinate, len(reviewer_status[0]) - 1))
    try:
        footers.append(create_formatting_regions(response_status_anchor.offset(row=len(response_status) - 1).coordinate, len(response_status[0]) - 1))
    except Exception as e:
        logger(f"Couldn't format the Response Status region header; perhaps there isn't a Response region? Error: {e}") 
    for range_string in footers:
        apply_styles_to_region(stat_footer_region_styles, range_string, ws_stats)



    wb.save()

    # if _WRITE_FILE:
    #     save_name = f'./dev/test/out/test_{timestamp()}.xlsx'
    #     wb.save(save_name)
    #     print(f'File {save_name} written to disk.')
    #     logger.debug(f'_WRITE_FILE = {_WRITE_FILE} -> saved workbook to "{save_name}"')

    # wb.close()


# if __name__ == '__main__':
#     main()