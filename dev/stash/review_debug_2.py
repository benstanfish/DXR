from dxbuild.dxreview import Review
from dxbuild.dxtools import timestamp, list_dimensions
import pandas as pd
import numpy as np
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

xml_path = './dev/test/data.xml'
review = Review.from_file(xml_path)
project_info = review.project_info
review_comments = review.review_comments

expansion_type = 'type'
all_list, rows = review_comments.to_list(expansion_type=expansion_type)
header_list, columns = review_comments._get_all_headers(expansion_type=expansion_type)


def copy_to_range(data_list: List, 
                  worksheet: Worksheet | None, 
                  anchor_cell: str='A1'):
    if worksheet is not None:
        anchor_row, anchor_column = coordinate_to_tuple(anchor_cell)
        rows, columns = list_dimensions(data_list)
        if rows == 1:
            for j in range(columns):
                worksheet.cell(row=anchor_row, column=anchor_column + j).value = data_list[j]
        else:
            for i in range(rows):
                for j in range(columns):
                    worksheet.cell(row=anchor_row + i, column=anchor_column + j).value = data_list[i][j]



xlsx_path = f'./dev/test/out/export_test_{timestamp()}.xlsx'

workbook = Workbook()
worksheet = workbook.active

HEADER_CELL = 'A2'
DATABODY_CELL = worksheet[HEADER_CELL].offset(1, 0).coordinate

# print(get_list_range(all_list, 'I11', to_transposed=True, return_R1C1=False))
# print(get_list_range(all_list, 'I11', to_transposed=True, return_R1C1=True))

anchor_row, anchor_column = coordinate_to_tuple(HEADER_CELL)
rows, columns = list_dimensions(all_list)


copy_to_range(header_list, worksheet=worksheet, anchor_cell=HEADER_CELL)
copy_to_range(all_list, worksheet=worksheet, anchor_cell=DATABODY_CELL)

# table_range = CellRange(min_row=anchor_row,
#                         min_col=anchor_column,
#                         max_row=anchor_row+columns - 1,
#                         max_col=anchor_column+rows)

# print(table_range.coord)

# table_style = TableStyleInfo(name="TableStyleMedium9", 
#                              showRowStripes=False, 
#                              showFirstColumn=False, 
#                              showLastColumn=False,
#                              showColumnStripes=False)

# table = Table(displayName='Comments',
#               ref=table_range.coord)

# table.tableStyleInfo = table_style
# if ws is not None:
#     ws.add_table(table)

workbook.save(xlsx_path)