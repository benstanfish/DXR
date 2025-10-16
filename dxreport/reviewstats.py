# Copyright (c) 2018-2025 Ben Fisher

import os
import datetime


from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import DEFAULT_FONT, Font, Alignment
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter

from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.legend import Legend
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.colors import ColorChoice


from PyQt6.QtWidgets import QApplication, QFileDialog

from dxbuild.reviews import Review
from dxbuild.constants import FALLBACKS, _LOG_DIR
from dxbuild.buildtools import timestamp, copy_to_range, apply_styles_to_region
from dxcore.cellformats import *
from dxreport import singlereport

if not os.path.exists(_LOG_DIR):
    os.makedirs(_LOG_DIR)

import logging
from dxcore.logconstants import log_format_string
logger = logging.getLogger(__name__)
logger.setLevel(logging.WARNING)
log_formatter = logging.Formatter(log_format_string)
log_file_handler = logging.FileHandler(f'{_LOG_DIR}/{__name__}.log')
log_file_handler.setFormatter(log_formatter)
logger.addHandler(log_file_handler)

_PADDING_OFFSET = 1


def make_stats_sheet(review: Review, ws: Worksheet) -> None:

    wb = ws.parent

    ws_table_name = ws.tables.items()[0][0]
    ws_stat_name = ws.title[:25] + '-STAT'

    ws_stats = wb.create_sheet(ws_stat_name)
    ws_stats.sheet_view.showGridLines = False
    
    comments = review.review_comments.comments

    disciplines = list(set([comment.discipline for comment in comments]))
    disciplines.sort()

    reviewers = list(set([comment.author for comment in comments]))
    reviewers.sort()

    highest_responses = list(set([comment.highest_response() for comment in comments]))
    highest_responses.sort()

    reviewers_with_open_comments = list(set([comment.author for comment in comments if comment.status == 'Open']))
    reviewers_with_open_comments.sort()
    
    reviewer_open_comments_dict = {}
    for reviewer in reviewers_with_open_comments:
        comment_ids = list(set([comment.id for comment in comments if (comment.status == 'Open' and comment.author == reviewer)]))
        reviewer_open_comments_dict[reviewer] = comment_ids


    xml_date = datetime.datetime.strptime(review.project_info.xml_date, '%Y-%m-%d %H:%M:%S')
    staleness = datetime.datetime.now() - xml_date
    
    recommend_message = ''
    if staleness.days > 7:
        recommend_message = ' The data is over 7 days old, downloading a new XML report is RECOMMENDED.'
    
    project_header_anchor = CellRange('A1')
    project_header = [
        ['Dr Checks Review Statistics',''],
        [f'This statistics run was created based on data exported on {review.project_info.xml_date}.', ''],
        [f'The age of this data is {staleness}.{recommend_message}', ''],
        ['', ''],
        ['Project Name', review.project_info.project_name],
        ['Project ID', review.project_info.project_id],
        ['Review Name', review.project_info.review_name],
    ]
    copy_to_range(project_header, ws_stats, project_header_anchor.coord)
    

    overall_status_anchor = ws_stats[project_header_anchor.coord].offset(row=len(project_header) + _PADDING_OFFSET, column=0)
    overall_status = [
        ['Overall Comment Status', '', '', ''],
        ['By Discipline', 'Open', 'Closed', 'Total']
    ]
    start_row = CellRange(overall_status_anchor.coordinate).min_row + len(overall_status)
    start_column = CellRange(overall_status_anchor.coordinate).min_col
    for i, discipline in enumerate(disciplines):
        temp_row = []
        temp_row.append(discipline)
        temp_row.append(f'=COUNTIFS({ws_table_name}[Discipline],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 1)}${start_row - 1})')
        temp_row.append(f'=COUNTIFS({ws_table_name}[Discipline],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 2)}${start_row - 1})')
        temp_row.append(f'=SUM({get_column_letter(start_column + 1)}{start_row + i}:{get_column_letter(start_column + 2)}{start_row + i})')
        overall_status.append(temp_row)
    temp_totals_row = ['Grand Total']
    for i in range(1, 4):
        temp_totals_row.append(f'=SUM({get_column_letter(start_column + i)}{start_row}:{get_column_letter(start_column + i)}{start_row + len(disciplines) - 1})')
    overall_status.append(temp_totals_row)
    copy_to_range(overall_status, ws_stats, overall_status_anchor.coordinate)

    
    reviewer_status_anchor = ws_stats[project_header_anchor.coord].offset(row=len(project_header) + _PADDING_OFFSET + 1, column=4 + _PADDING_OFFSET)
    reviewer_status = [
        ['By Author', 'Open', 'Closed', 'Total']
    ]
    start_row = CellRange(reviewer_status_anchor.coordinate).min_row + len(reviewer_status)
    start_column = CellRange(reviewer_status_anchor.coordinate).min_col    
    for i, reviewer in enumerate(reviewers):
        temp_row = []
        temp_row.append(reviewer)
        temp_row.append(f'=COUNTIFS({ws_table_name}[Author],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 1)}${start_row - 1})')
        temp_row.append(f'=COUNTIFS({ws_table_name}[Author],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 2)}${start_row - 1})')
        temp_row.append(f'=SUM({get_column_letter(start_column + 1)}{start_row + i}:{get_column_letter(start_column + 2)}{start_row + i})')
        reviewer_status.append(temp_row)
    temp_totals_row = ['Grand Total']
    for i in range(1, 4):
        temp_totals_row.append(f'=SUM({get_column_letter(start_column + i)}{start_row}:{get_column_letter(start_column + i)}{start_row + len(reviewers) - 1})')
    reviewer_status.append(temp_totals_row)
    copy_to_range(reviewer_status, ws_stats, reviewer_status_anchor.coordinate)    


    try:
        response_status_anchor = reviewer_status_anchor.offset(row=0, column=4 + _PADDING_OFFSET)
        response_status = [
            ['By Response', 'Open', 'Closed', 'Total']
        ]
        start_row = CellRange(response_status_anchor.coordinate).min_row + len(response_status)
        start_column = CellRange(response_status_anchor.coordinate).min_col  
        for i, highest_response in enumerate(highest_responses):
            temp_row = []
            if highest_response == '':
                temp_row.append('None')
                temp_row.append(f'=COUNTIFS({ws_table_name}[Highest Resp.],"",{ws_table_name}[Status],{get_column_letter(start_column + 1)}${start_row - 1})')
                temp_row.append(f'=COUNTIFS({ws_table_name}[Highest Resp.],"",{ws_table_name}[Status],{get_column_letter(start_column + 2)}${start_row - 1})')
            else:
                temp_row.append(highest_response)
                temp_row.append(f'=COUNTIFS({ws_table_name}[Highest Resp.],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 1)}${start_row - 1})')
                temp_row.append(f'=COUNTIFS({ws_table_name}[Highest Resp.],${get_column_letter(start_column)}{start_row + i},{ws_table_name}[Status],{get_column_letter(start_column + 2)}${start_row - 1})')
            temp_row.append(f'=SUM({get_column_letter(start_column + 1)}{start_row + i}:{get_column_letter(start_column + 2)}{start_row + i})')
            response_status.append(temp_row)
        temp_totals_row = ['Grand Total']
        for i in range(1, 4):
            temp_totals_row.append(f'=SUM({get_column_letter(start_column + i)}{start_row}:{get_column_letter(start_column + i)}{start_row + len(highest_responses) - 1})')
        response_status.append(temp_totals_row)
        copy_to_range(response_status, ws_stats, response_status_anchor.coordinate)   
    except Exception as e:
        print(f"Couldn't write the response region; likely no responses yet. Error at {e}.")
        logger.exception(f"Couldn't write the response region; likely no responses yet. Error at {e}.")


    status_region_max_rows = max(len(overall_status), len(reviewer_status))

    open_comments_anchor = ws_stats[project_header_anchor.coord].offset(row=len(project_header) + _PADDING_OFFSET + status_region_max_rows + _PADDING_OFFSET + 1, column=0)
    open_comments_anchor.value = 'Open Comments by Author'
    for i, reviewer in enumerate(reviewer_open_comments_dict.keys()):
        open_comments_anchor.offset(row=1, column=i).value = reviewer
        comment_id_list = reviewer_open_comments_dict[reviewer]
        comment_id_list.sort()
        for j, comment_id in enumerate(comment_id_list):
            open_comments_anchor.offset(row=2 + j, column=i).value = comment_id


    max_open_comments = 0
    for key in reviewer_open_comments_dict.keys():
        if len(reviewer_open_comments_dict[key]) > max_open_comments:
            max_open_comments = len(reviewer_open_comments_dict[key])

    review_open_comments_region_height = max_open_comments + 2
    
    
    
    open_by_discipline_evaluator = list(set([comment.discipline for comment in review.review_comments.comments if comment.status == 'Open' and comment.ball_in_court == 'Evaluator']))
    open_by_discipline_evaluator.sort()   
    bic_evaluator_dic = {}
    for discipline in open_by_discipline_evaluator:
        bic_evaluator_dic[discipline] = list(set([comment.id for comment in review.review_comments.comments if comment.discipline == discipline and comment.status == 'Open' and comment.ball_in_court == 'Evaluator']))

    ball_in_court_evaluator_anchor = ws_stats[project_header_anchor.coord].offset(
        row=len(project_header) + _PADDING_OFFSET + status_region_max_rows + _PADDING_OFFSET + review_open_comments_region_height + _PADDING_OFFSET + 1, column=0)
    ball_in_court_evaluator_anchor.value = 'Open Comments --- Ball in Court: Evaluator'
    for i, discipline in enumerate(bic_evaluator_dic.keys()):
        ball_in_court_evaluator_anchor.offset(row=1, column=i).value = discipline
        comment_id_list = bic_evaluator_dic[discipline]
        comment_id_list.sort()
        for j, comment_id in enumerate(comment_id_list):
            ball_in_court_evaluator_anchor.offset(row=2 + j, column=i).value = comment_id    
    
    
    max_bic_evaluator_comments = 0
    for key in bic_evaluator_dic.keys():
        if len(bic_evaluator_dic[key]) > max_bic_evaluator_comments:
            max_bic_evaluator_comments = len(bic_evaluator_dic[key])
            
    bic_evaluator_comments_region_height = max_bic_evaluator_comments + 2
        
    open_by_discipline_commentor = list(set([comment.discipline for comment in review.review_comments.comments if comment.status == 'Open' and comment.ball_in_court == 'Commentor']))
    open_by_discipline_commentor.sort()
    bic_commentor_dic = {}
    for discipline in open_by_discipline_commentor:
        bic_commentor_dic[discipline] = list(set([comment.id for comment in review.review_comments.comments if comment.discipline == discipline and comment.status == 'Open' and comment.ball_in_court == 'Commentor']))
       
    ball_in_court_commentor_anchor = ws_stats[project_header_anchor.coord].offset(
        row=len(project_header) + _PADDING_OFFSET + 
            status_region_max_rows + _PADDING_OFFSET + 
            review_open_comments_region_height + _PADDING_OFFSET +
            bic_evaluator_comments_region_height + _PADDING_OFFSET + 1, column=0)
    ball_in_court_commentor_anchor.value = 'Open Comments --- Ball in Court: Commentor'
    for i, discipline in enumerate(bic_commentor_dic.keys()):
        ball_in_court_commentor_anchor.offset(row=1, column=i).value = discipline
        comment_id_list = bic_commentor_dic[discipline]
        comment_id_list.sort()
        for j, comment_id in enumerate(comment_id_list):
            ball_in_court_commentor_anchor.offset(row=2 + j, column=i).value = comment_id    
    


    # Global formatting: reset
    for row in ws_stats.iter_rows():
        for cell in row:
            cell.font = Font(name='Aptos Narrow', sz=10)
            cell.alignment = Alignment(horizontal='left')

    # Lower priority column width changes first
    for i, _ in enumerate(reviewer_open_comments_dict.keys()):
        ws_stats.column_dimensions[get_column_letter(open_comments_anchor.offset(column=i).column)].width = Widths.SMALL_12

    # Format the title and region titles
    ws_stats[project_header_anchor.coord].font = Font(name='Aptos', sz=14, bold=True)
    ws_stats[overall_status_anchor.coordinate].font = Font(name='Aptos', sz=12, bold=True)
    ws_stats[open_comments_anchor.coordinate].font = Font(name='Aptos', sz=12, bold=True)
    ws_stats[ball_in_court_evaluator_anchor.coordinate].font = Font(name='Aptos', sz=12, bold=True)
    ws_stats[ball_in_court_commentor_anchor.coordinate].font = Font(name='Aptos', sz=12, bold=True)

    ws_stats.column_dimensions[get_column_letter(overall_status_anchor.column)].width = Widths.SMALL_12
    ws_stats.column_dimensions[get_column_letter(reviewer_status_anchor.column)].width = Widths.SMALL_12
    try:
        ws_stats.column_dimensions[get_column_letter(response_status_anchor.column)].width = Widths.SMALL_12
    except Exception as e:
        logger(f"Couldn't change the width of the Response Status region; perhaps there isn't a Response region? Error: {e}")


    # Format the bold fields in the project header region
    for i in range(4, len(project_header)):
        ws_stats[project_header_anchor.coord].offset(row=i, column=0).font = Font(name='Aptos Narrow', sz=10, bold=True)
        if i == 4:
            ws_stats[project_header_anchor.coord].offset(row=i, column=1).font = Font(name='Aptos Narrow', sz=12, bold=True)


    # Format the headers of each overall regions
    def create_formatting_regions(initial_anchor: str, region_width: int) -> str:
        temp_range = CellRange(range_string=initial_anchor)
        temp_range.expand(right=region_width)
        return temp_range.coord

    headers = []
    headers.append(create_formatting_regions(overall_status_anchor.offset(row=1).coordinate, len(overall_status[0]) - 1))
    headers.append(create_formatting_regions(reviewer_status_anchor.coordinate, len(reviewer_status[0]) - 1))
    try:
        headers.append(create_formatting_regions(response_status_anchor.coordinate, len(response_status[0]) - 1))
    except Exception as e:
        logger(f"Couldn't format the Response Status region header; perhaps there isn't a Response region? Error: {e}") 

    headers.append(create_formatting_regions(open_comments_anchor.offset(row=1).coordinate, len(reviewer_open_comments_dict.keys()) - 1))
    if len(open_by_discipline_evaluator) > 0:
        headers.append(create_formatting_regions(ball_in_court_evaluator_anchor.offset(row=1).coordinate, len(bic_evaluator_dic.keys()) - 1))
    else:
        ball_in_court_evaluator_anchor.offset(row=1).value = "None"
    if len(open_by_discipline_commentor) > 0:
        headers.append(create_formatting_regions(ball_in_court_commentor_anchor.offset(row=1).coordinate, len(bic_commentor_dic.keys()) - 1))
    else:
        ball_in_court_commentor_anchor.offset(row=1).value = "None"
    for range_string in headers:
        apply_styles_to_region(stat_header_region_styles, range_string, ws_stats)    
    
    
    footers = []
    footers.append(create_formatting_regions(overall_status_anchor.offset(row=len(overall_status) - 1).coordinate, len(overall_status[0]) - 1))
    footers.append(create_formatting_regions(reviewer_status_anchor.offset(row=len(reviewer_status) - 1).coordinate, len(reviewer_status[0]) - 1))
    try:
        footers.append(create_formatting_regions(response_status_anchor.offset(row=len(response_status) - 1).coordinate, len(response_status[0]) - 1))
    except Exception as e:
        logger(f"Couldn't format the Response Status region header; perhaps there isn't a Response region? Error: {e}") 
    for range_string in footers:
        apply_styles_to_region(stat_footer_region_styles, range_string, ws_stats)



    
    
    # Add ChartObjects
    # Comment Count by Discipline
    disc_chart = BarChart()
    disc_chart.type = 'bar'
    disc_chart.grouping = 'stacked'
    disc_chart.overlap = 100
    disc_chart.title = 'Comments by Discipline'

    disc_chart.graphical_properties = GraphicalProperties()
    disc_chart.graphical_properties.line.noFill = True

    disc_chart.height = 8
    disc_chart.width = 14

    disc_chart.layout = Layout(manualLayout=ManualLayout(xMode='edge',
                                                         yMode='edge',
                                                         x=0,
                                                         y=0.2,
                                                         h=0.8, 
                                                         w=1))
    disc_chart.layout.layoutTarget = 'inner'

    disc_chart.legend = Legend()
    disc_chart.legend.layout = Layout(manualLayout=ManualLayout(xMode='edge',
                                                                yMode='edge',
                                                                x=0.375,
                                                                y=0.1,
                                                                h=0.1, 
                                                                w=0.25))

    disc_chart.x_axis.delete = False
    disc_chart.x_axis.graphicalProperties = GraphicalProperties()
    disc_chart.x_axis.graphicalProperties.line.solidFill = WebColor.DARKSLATEGRAY
    
    disc_chart.y_axis.delete = False
    disc_chart.y_axis.graphicalProperties = GraphicalProperties()
    disc_chart.y_axis.graphicalProperties.line.noFill = True

    disc_chart.y_axis.majorGridlines = ChartLines()
    disc_chart.y_axis.majorGridlines.spPr = GraphicalProperties()
    disc_chart.y_axis.majorGridlines.spPr.line.solidFill = WebColor.LIGHTGRAY

    categories_cell_range = CellRange(range_string=f'{overall_status_anchor.offset(row=2, column=0).coordinate}:' \
                                                   f'{overall_status_anchor.offset(row=2 + len(disciplines) - 1, column=0).coordinate}')
    categories = Reference(worksheet=ws_stats, 
                          min_col=categories_cell_range.min_col, 
                          max_col=categories_cell_range.max_col,
                          min_row=categories_cell_range.min_row,
                          max_row=categories_cell_range.max_row
                          )
    
    open_cell_range = CellRange(range_string=f'{overall_status_anchor.offset(row=2, column=1).coordinate}:' \
                                             f'{overall_status_anchor.offset(row=2 + len(disciplines) - 1, column=1).coordinate}')
    open_data = Reference(worksheet=ws_stats, 
                        min_col=open_cell_range.min_col, 
                        max_col=open_cell_range.max_col,
                        min_row=open_cell_range.min_row,
                        max_row=open_cell_range.max_row
                        )
    open_series = Series(open_data, title='Open')
    open_series.graphicalProperties.solidFill = ColorChoice(srgbClr=WebColor.TOMATO)
    disc_chart.append(open_series)
    
    closed_cell_range =  CellRange(range_string=f'{overall_status_anchor.offset(row=2, column=2).coordinate}:' \
                                                f'{overall_status_anchor.offset(row=2 + len(disciplines) - 1, column=2).coordinate}')
    closed_data = Reference(worksheet=ws_stats, 
                        min_col=closed_cell_range.min_col, 
                        max_col=closed_cell_range.max_col,
                        min_row=closed_cell_range.min_row,
                        max_row=closed_cell_range.max_row
                        )
    closed_series = Series(closed_data, title='Closed')
    closed_series.graphicalProperties.solidFill = ColorChoice(srgbClr=WebColor.DODGERBLUE)
    disc_chart.append(closed_series)

    disc_chart.set_categories(categories)
    ws_stats.add_chart(disc_chart, 'P5')



    # Comment Count by Author
    auth_chart = BarChart()
    auth_chart.type = 'bar'
    auth_chart.grouping = 'stacked'
    auth_chart.overlap = 100
    auth_chart.title = 'Comments by Author'

    auth_chart.graphical_properties = GraphicalProperties()
    auth_chart.graphical_properties.line.noFill = True

    auth_chart.height = 8
    auth_chart.width = 14

    auth_chart.layout = Layout(manualLayout=ManualLayout(xMode='edge',
                                                         yMode='edge',
                                                         x=0,
                                                         y=0.2,
                                                         h=0.8, 
                                                         w=1))
    auth_chart.layout.layoutTarget = 'inner'

    auth_chart.legend = Legend()
    auth_chart.legend.layout = Layout(manualLayout=ManualLayout(xMode='edge',
                                                                yMode='edge',
                                                                x=0.375,
                                                                y=0.1,
                                                                h=0.1, 
                                                                w=0.25))

    auth_chart.x_axis.delete = False
    auth_chart.x_axis.graphicalProperties = GraphicalProperties()
    auth_chart.x_axis.graphicalProperties.line.solidFill = WebColor.DARKSLATEGRAY
    
    auth_chart.y_axis.delete = False
    auth_chart.y_axis.graphicalProperties = GraphicalProperties()
    auth_chart.y_axis.graphicalProperties.line.noFill = True

    auth_chart.y_axis.majorGridlines = ChartLines()
    auth_chart.y_axis.majorGridlines.spPr = GraphicalProperties()
    auth_chart.y_axis.majorGridlines.spPr.line.solidFill = WebColor.LIGHTGRAY

    auth_categories_cell_range = CellRange(range_string=f'{reviewer_status_anchor.offset(row=1, column=0).coordinate}:' \
                                                   f'{reviewer_status_anchor.offset(row=1 + len(reviewers) - 1, column=0).coordinate}')
    auth_categories = Reference(worksheet=ws_stats, 
                          min_col=auth_categories_cell_range.min_col, 
                          max_col=auth_categories_cell_range.max_col,
                          min_row=auth_categories_cell_range.min_row,
                          max_row=auth_categories_cell_range.max_row
                          )
    
    auth_open_cell_range = CellRange(range_string=f'{reviewer_status_anchor.offset(row=1, column=1).coordinate}:' \
                                             f'{reviewer_status_anchor.offset(row=1 + len(reviewers) - 1, column=1).coordinate}')
    
    auth_open_data = Reference(worksheet=ws_stats, 
                        min_col=auth_open_cell_range.min_col, 
                        max_col=auth_open_cell_range.max_col,
                        min_row=auth_open_cell_range.min_row,
                        max_row=auth_open_cell_range.max_row
                        )
    auth_open_series = Series(auth_open_data, title='Open')
    auth_open_series.graphicalProperties.solidFill = ColorChoice(srgbClr=WebColor.TOMATO)
    auth_chart.append(auth_open_series)
    
    auth_closed_cell_range =  CellRange(range_string=f'{reviewer_status_anchor.offset(row=1, column=2).coordinate}:' \
                                                f'{reviewer_status_anchor.offset(row=1 + len(reviewers) - 1, column=2).coordinate}')
    auth_closed_data = Reference(worksheet=ws_stats, 
                        min_col=auth_closed_cell_range.min_col, 
                        max_col=auth_closed_cell_range.max_col,
                        min_row=auth_closed_cell_range.min_row,
                        max_row=auth_closed_cell_range.max_row
                        )
    auth_closed_series = Series(auth_closed_data, title='Closed')
    auth_closed_series.graphicalProperties.solidFill = ColorChoice(srgbClr=WebColor.DODGERBLUE)
    auth_chart.append(auth_closed_series)

    auth_chart.set_categories(auth_categories)
    ws_stats.add_chart(auth_chart, 'Y5')



