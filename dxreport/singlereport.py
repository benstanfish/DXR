# Copyright (c) 2018-2025 Ben Fisher

from datetime import datetime

from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.styles.alignment import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from dxbuild.reviews import Review
import dxbuild.buildtools as buildtools
from dxbuild.constants import RESPONSE_COLUMNS, USER_NOTES_COLUMNS
from dxcore.conditionalformats import *
from dxcore.cellformats import *
from dxbuild.constants import FALLBACKS, _TRUE_SYMBOLIC


def create_report(ws: Worksheet, xml_path: str):

    review = Review.from_file(xml_path)
    project_info = review.project_info
    all_comments = review.review_comments
    user_notes = review.user_notes

    if ws is not None and review:
        
        ws.title = project_info.review_name[:30]

        # Get upper left cell of each major region
        user_notes_anchor_cell = review.user_notes.get_anchor_cell(ws, 'extents')
        project_info_anchor_cell = review.project_info.get_anchor_cell(ws, 'extents')
        review_comments_anchor = review.review_comments.get_anchor_cell(ws, 'extents')

        # Dump Data to Excel
        buildtools.copy_to_range(user_notes.to_list(), worksheet=ws, anchor_cell=user_notes_anchor_cell)
        buildtools.copy_to_range(project_info.to_list(), worksheet=ws, anchor_cell=project_info_anchor_cell)
        buildtools.copy_to_range(all_comments.to_list(), worksheet=ws, anchor_cell=review_comments_anchor)

        user_notes.autonumber_id_column(ws)

        TABLE_RANGE = review.frames['extents'].coord
        TABLE_HEADER = review.frames['header'].coord
        TABLE_BODY = review.frames['body'].coord

        table_names = []
        wb = ws.parent
        for sht in wb.worksheets:
            for table in sht.tables:
                table_names.append(table.name)

        new_table_name = buildtools.autoincrement_name(base_name='Comments', search_list=table_names)
        table = Table(displayName=new_table_name, ref=TABLE_RANGE)

        ws.add_table(table)
        ws.sheet_view.showGridLines = False
        TABLE_INFO = buildtools.get_table_info(ws)
        
        review.build_table_column_list(ws)
        table_column_list = review.table_column_list

        top_left_alignment = Alignment(horizontal='left', vertical='top')
        used_range = ws.calculate_dimension()
        for row in ws[used_range]:
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'm/d/yy'
                cell.alignment = top_left_alignment

        buildtools.apply_styles_to_region(project_info_key_style, review.project_info.frames['keys'].coord, ws)
        buildtools.apply_styles_to_region(project_title_style, review.project_info.frames['project_title'].coord, ws)


        buildtools.apply_styles_to_region(table_header_styles, TABLE_HEADER, ws)
        buildtools.apply_styles_to_region(table_body_styles, TABLE_BODY, ws)
        buildtools.apply_styles_to_region(user_notes_header_styles, review.user_notes.frames['header'].coord, ws)
        buildtools.apply_styles_to_region(user_notes_body_styles, review.user_notes.frames['body'].coord, ws)

        user_notes_proposed_statuses = 'Open, For Information Only, Non-Concur, Check and Resolve, Open, Closed, Closed without Comment'    
        user_notes_proposed_statuses_column = buildtools.build_column_vectors(buildtools.get_columns_by_name('Proposed State', TABLE_INFO), TABLE_INFO)
        buildtools.add_data_validation_to_column(user_notes_proposed_statuses, user_notes_proposed_statuses_column, ws)
        buildtools.conditionally_format_range(user_notes_proposed_statuses_column[0], 'concur', ws, light_blue_dx)    
        buildtools.conditionally_format_range(user_notes_proposed_statuses_column[0], 'for information only', ws, light_green_dx)
        buildtools.conditionally_format_range(user_notes_proposed_statuses_column[0], 'non-concur', ws, light_yellow_dx)
        buildtools.conditionally_format_range(user_notes_proposed_statuses_column[0], 'check and resolve', ws, light_red_dx)
        buildtools.conditionally_format_range(user_notes_proposed_statuses_column[0], 'open', ws, light_yellow_dx)
        buildtools.conditionally_format_range(user_notes_proposed_statuses_column[0], 'closed', ws, light_green_dx)
        buildtools.conditionally_format_range(user_notes_proposed_statuses_column[0], 'closed without comment', ws, light_blue_dx)

        user_notes_states = 'Working, Delegated, Ready, Done, N/A'
        user_notes_states_column = buildtools.build_column_vectors(buildtools.get_columns_by_name('State', TABLE_INFO), TABLE_INFO)
        buildtools.add_data_validation_to_column(user_notes_states, user_notes_states_column, ws)
        buildtools.conditionally_format_range(user_notes_states_column[0], 'working', ws, red_dx)
        buildtools.conditionally_format_range(user_notes_states_column[0], 'delegated', ws, yellow_dx)
        buildtools.conditionally_format_range(user_notes_states_column[0], 'ready', ws, green_dx)
        buildtools.conditionally_format_range(user_notes_states_column[0], 'done', ws, blue_dx)
        buildtools.conditionally_format_range(user_notes_states_column[0], 'n/a', ws, gray_dx)

        status_options = 'Concur, For Information Only, Non-Concur, Check and Resolve'
        status_column_letters = buildtools.get_columns_by_name('status', TABLE_INFO)
        status_columns = buildtools.build_column_vectors(status_column_letters, TABLE_INFO) # Treat the first occurance separately
        buildtools.add_data_validation_to_column(status_options, status_columns[1:], ws)

        open_closed_options = 'Open, Closed'
        first_status_column = buildtools.add_data_validation_to_column(open_closed_options, status_columns[:1], ws)
        buildtools.conditionally_format_range(status_columns[0], 
                                            'closed', 
                                            ws, 
                                            light_gray_dx, 
                                            review.review_comments.frames['body'].coord, 
                                            stop_if_true=True)
            
        for column in status_columns[1:]:
            buildtools.conditionally_format_range(column, 'check and resolve', ws, light_red_dx)
            buildtools.conditionally_format_range(column, 'non-concur', ws, light_yellow_dx)
            buildtools.conditionally_format_range(column, 'for information only', ws, light_green_dx)
            buildtools.conditionally_format_range(column, 'concur', ws, light_blue_dx)
            buildtools.apply_styles_to_region_if_empty(empty_status_cell_style, column, ws)
        
        highest_reponse_letters = buildtools.get_columns_by_name('highest resp', TABLE_INFO)
        highest_response_columns = buildtools.build_column_vectors(highest_reponse_letters, TABLE_INFO)[0]
        buildtools.add_data_validation_to_column(status_options, [highest_response_columns], ws)
        buildtools.conditionally_format_range(highest_response_columns, 'check and resolve', ws, red_dx)
        buildtools.conditionally_format_range(highest_response_columns, 'non-concur', ws, yellow_dx)
        buildtools.conditionally_format_range(highest_response_columns, 'for information only', ws, green_dx)
        buildtools.conditionally_format_range(highest_response_columns, 'concur', ws, blue_dx)
        buildtools.apply_styles_to_region_if_empty(empty_status_cell_style, highest_response_columns, ws)

        critical_options = 'Yes, No'
        critical_column_letters = buildtools.get_columns_by_name('critical', TABLE_INFO)
        critical_column = buildtools.build_column_vectors(critical_column_letters, TABLE_INFO)
        buildtools.add_data_validation_to_column(critical_options, critical_column, ws)
        buildtools.conditionally_format_range(critical_column[0], 'yes', ws, red_dx)    
        
        class_options = 'CUI, Unclassified, Public'
        class_column_letters = buildtools.get_columns_by_name('class', TABLE_INFO)
        class_column = buildtools.build_column_vectors(class_column_letters, TABLE_INFO)
        buildtools.add_data_validation_to_column(class_options, class_column, ws)
        buildtools.conditionally_format_range(class_column[0], 'cui', ws, red_dx)   
        buildtools.conditionally_format_range(class_column[0], 'unclassified', ws, yellow_dx)

        att_options = f'{_TRUE_SYMBOLIC}, '
        att_column_letters = buildtools.get_columns_by_name('att', TABLE_INFO)
        att_columns = buildtools.build_column_vectors(att_column_letters, TABLE_INFO)    
        for column in att_columns:
            buildtools.add_data_validation_to_column(att_options, [column], ws)
            buildtools.conditionally_format_range(column, _TRUE_SYMBOLIC, ws, has_att_dx)  
        
        for i, col_width in enumerate(USER_NOTES_WIDTHS):
            ws.column_dimensions[get_column_letter(i + 1)].width = col_width

        for i, col_width in enumerate(COMMENT_COLUMN_WIDTHS):
            ws.column_dimensions[get_column_letter(i + review.user_notes.count + 1)].width = col_width

        if review.review_comments.responses_count > 0:
            response_header_cell_range = review.review_comments.frames['response_header']
            for i in range(response_header_cell_range.min_col, response_header_cell_range.max_col + 1):
                i_modulo = (i - response_header_cell_range.min_col) % len(RESPONSE_COLUMNS)
                ws.column_dimensions[f'{get_column_letter(i)}'].width = RESPONSE_COLUMN_WIDTHS[i_modulo]

        wrappables = USER_NOTES_COLUMNS + ['Comment']
        #NOTE: add 'Text' to the list above to wrap the Text column in the responses. Right now
        #it's not wrapping because Openpyxl column_dimensions.group() changes reverts the column widths
        #and I need to figure out how to prevent that from happening.
        for wrappable in wrappables:
            for item in review.table_column_list:
                if wrappable in item:
                    range_string = f'{get_column_letter(review.table_column_list.index(item) + 1)}{review.frames['body'].min_row}:' \
                                f'{get_column_letter(review.table_column_list.index(item) + 1)}{review.frames['body'].max_row}'
                    buildtools.apply_styles_to_region(table_body_wrap_styles, range_string, ws)

        # Group columns
        _COLLAPSE_LEFT = '▷'
        _COLLAPSE_RIGHT = '◁ Expand'
        collapse_regions = [['Notes', 'State'],
                            ['Source', 'Section']]

        if review.review_comments.responses_count > 0:
            response_collapse_region = []   # This is used in conjunction wiht the reponse_header_cell_range from above
            for col in range(response_header_cell_range.min_col, response_header_cell_range.max_col + 1):
                current_modulo = (col - response_header_cell_range.min_col) % len(RESPONSE_COLUMNS)
                if current_modulo == 1 or current_modulo == len(RESPONSE_COLUMNS) - 1:
                    response_collapse_region.append(table_column_list[col - 1])
                if current_modulo == len(RESPONSE_COLUMNS) - 1:
                    collapse_regions.append(response_collapse_region)
                    response_collapse_region = []
            
        for region in collapse_regions:
            start_col = table_column_list.index(region[0]) + 1
            stop_col = table_column_list.index(region[1]) + 1
            
            label_left_cell = ws[f'{get_column_letter(start_col - 1)}{review.frames['header'].min_row - 1}']
            label_left_cell.value = _COLLAPSE_LEFT
            label_left_cell.alignment = Alignment(horizontal='right', vertical='top')

            #NOTE: There is an Openpyxl bug that resets column widths when grouping is used. The workaround is to
            # set the width of the first column in the group AFTER grouping, however the subsequent columns in the
            # group inherent the same width... ugh.
            ws.column_dimensions.group(get_column_letter(start_col), 
                                    get_column_letter(stop_col), 
                                    hidden=True)

        # This loop is purely needed to make sure the _COLLAPSE_RIGHT label overwrites _COLLAPSE_LEFT at collisions
        for region in collapse_regions:
            start_col = table_column_list.index(region[0]) + 1
            stop_col = table_column_list.index(region[1]) + 1
            
            label_right_cell = ws[f'{get_column_letter(stop_col + 1)}{review.frames['header'].min_row - 1}']
            label_right_cell.value = _COLLAPSE_RIGHT


        return True
    return False