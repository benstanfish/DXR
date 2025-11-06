# Copyright (c) 2018-2025 Ben Fisher

import sys, os, subprocess

from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT
from openpyxl.worksheet.worksheet import Worksheet

from PyQt6.QtWidgets import QApplication, QFileDialog

from dxbuild.reviews import Review
from dxbuild.variables import FALLBACKS
from dxbuild.buildtools import timestamp, clean_name, autoincrement_name
from dxreport import singlereport, reviewstats

# import logging
# from constants import LOG_DIR
# from dxcore.logconstants import log_format_string
# logger = logging.getLogger(__name__)
# logger.setLevel(logging.WARNING)
# log_formatter = logging.Formatter(log_format_string)
# log_file_handler = logging.FileHandler(f'{LOG_DIR}/{__name__}.log')
# log_file_handler.setFormatter(log_formatter)
# logger.addHandler(log_file_handler)

# _DEBUG_MODE = False
# test_path = './dev/test/data.xml'

def batch_create_reports() -> str | bool:

    _WRITE_FILE = True

    DEFAULT_FONT.__init__(name=FALLBACKS['font_name'], size=FALLBACKS['font_size'])
    wb = Workbook()
    
    xml_paths, _ = QFileDialog.getOpenFileNames(
        parent=None, 
        caption='Select Files Dialog',
        filter='XML Files (*.xml);;HTML Files (*.html);;All Files (*)'
    )

    # Check to see if the user selected any files or cancelled the file select dialog.
    if len(xml_paths) > 0:
        for i, xml_path in enumerate(xml_paths):
            review = Review.from_file(xml_path)
            # logger.debug(f'Created review from {xml_path} and review.is_valid returned {review.is_valid}')
            if review:
                if len(wb.sheetnames) == 1 and i == 0: 
                    
                    ws = wb.active
                    ws.title = clean_name(review.project_info.review_name[:30])
                else:
                    next_name = autoincrement_name(review.project_info.review_name[:30], wb.sheetnames, True)
                    ws = wb.create_sheet(next_name)
                # logger.debug(f'Acquired worksheet {i + 1}, "{ws.title}", to insert report from "{xml_path}"')
                singlereport.create_report(review, ws)
                reviewstats.make_stats_sheet(review, ws)
                                
        if _WRITE_FILE:
            save_name = os.path.join(os.path.dirname(xml_paths[0]), f'DrChecks Summary Report {timestamp('%Y-%m-%d %H-%M-%S')}.xlsx')
            # save_name = os.path.join(os.path.dirname(xml_paths[0]), f'DrChecks Summary Report {timestamp('%Y-%m-%d %H-%M-%S')}.xlsx')
            wb.save(save_name)
            # logger.debug(f'_WRITE_FILE = {_WRITE_FILE} -> saved workbook to {save_name}')
        wb.close()
        print(f'{save_name} written to disk.')
        os.startfile(os.path.dirname(save_name))
        os.startfile(save_name)
        # subprocess.Popen(f'explorer /select,"{save_name}"') 
        return save_name
    else:
        # logger.debug('File dialog closed without selecting files.')
        print('There was an error, refer to the logs.')
        return False

# if __name__ == '__main__':
#     batch_create_reports()