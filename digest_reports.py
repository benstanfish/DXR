# Copyright (c) 2018-2025 Ben Fisher

import sys, os
if not os.path.exists('./logs'):
    os.makedirs('./logs')

from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT
from openpyxl.worksheet.worksheet import Worksheet

from PyQt6.QtWidgets import QApplication, QFileDialog

from dxbuild.reviews import Review
from dxbuild.constants import FALLBACKS
from dxbuild.buildtools import timestamp, is_valid_root
from dxbuild.buildtools import autoincrement_name
from dxreport import singlereport

import logging
from dxcore.logconstants import log_format_string
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
log_formatter = logging.Formatter(log_format_string)
log_file_handler = logging.FileHandler(f'./logs/{__name__}.log')
log_file_handler.setFormatter(log_formatter)
logger.addHandler(log_file_handler)


def create_next_worksheet(ws_basename: str, workbook: Workbook) -> Worksheet:
    """Add an autoincremented worksheet to existing workbook."""
    print('entered')
    # modified_name = autoincrement_name(ws_basename, workbook.sheetnames)
    modified_name = ws_basename
    if len(workbook.sheetnames) == 1:
        print('in if')
        ws = workbook.active
        ws.title = modified_name
    else:
        print('in else')
        ws = workbook.create_sheet(modified_name)
    print(type(ws))
    return ws
    

def main() -> None:

    _WRITE_FILE = True
    xml_path = './dev/test/data.xml'

    DEFAULT_FONT.__init__(name=FALLBACKS['font_name'], size=FALLBACKS['font_size'])
    wb = Workbook()

    app = QApplication(sys.argv)
    xml_paths, _ = QFileDialog.getOpenFileNames(
        parent=None, 
        caption='Select Files Dialog',
        filter='XML Files (*.xml);;HTML Files (*.html);;All Files (*)'
    )

    # Check to see if the user selected any files or cancelled the file select dialog.
    if len(xml_paths) > 0:

        for xml_path in xml_paths:
            # if is_valid_root(xml_path=xml_path):
            #     logger.debug(f'will attempt to validate the root for {xml_path}')
            #     try:
            #         review = Review.from_file(xml_path)
            #         logger.debug(f'File path: {xml_path} returns {review.is_valid} for a valid root "ProjNet".')
            #         #TODO: this is where the ws needs to be created and review added
            #         if review:
            #             # try:
            #             ws = create_next_worksheet(ws_basename=review.project_info.review_name[30:], workbook=wb)
            #             singlereport.create_report(review, ws)
            #             logger.info(f'single report was ran.')
            #             # except Exception as e:
            #             #     logger.exception(f'File path: {xml_path} resulted in an error: {e}') 

            #     except Exception as e:
            #         logger.exception(f'File path: {xml_path} resulted in an error: {e}')

            review = Review.from_file(xml_path)
            
            if len(wb.sheetnames) == 1:
                ws = wb.active
                ws.title = review.project_info.review_name[:30]
            else:
                ws = wb.create_sheet(review.project_info.review_name[:30])

            
            if review:
                singlereport.create_report(review, ws)

        if _WRITE_FILE:
            save_name = f'./dev/test/out/test_{timestamp()}.xlsx'
            wb.save(save_name)
            print(f'{save_name} has been saved.')

        wb.close()

    else:
        logger.debug('File dialog closed without selecting files.')

if __name__ == '__main__':
    main()