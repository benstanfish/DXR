# Copyright (c) 2018-2025 Ben Fisher

from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from .dxcolor import DXColor
from dxbuild.constants import FALLBACKS

_FONT_NAME = FALLBACKS['font_name']
_FONT_SIZE = FALLBACKS['font_size']


red_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=True, color=DXColor.WHITE), 
    fill=PatternFill(start_color=DXColor.RED, end_color=DXColor.RED)
)

yellow_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=False, color=DXColor.BLACK), 
    fill=PatternFill(start_color=DXColor.YELLOW, end_color=DXColor.YELLOW)
)

green_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=False, color=DXColor.WHITE), 
    fill=PatternFill(start_color=DXColor.GREEN, end_color=DXColor.GREEN)
)

blue_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=False, color=DXColor.WHITE), 
    fill=PatternFill(start_color=DXColor.BLUE, end_color=DXColor.BLUE)
)

gray_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=False, color=DXColor.WHITE), 
    fill=PatternFill(start_color=DXColor.GRAY, end_color=DXColor.GRAY)
)

light_red_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=True, color=DXColor.RED_TEXT), 
    fill=PatternFill(start_color=DXColor.LIGHT_RED_BG, end_color=DXColor.LIGHT_RED_BG),
    border=Border(left=Side(border_style='thin', color=DXColor.RED_TEXT))
)

light_yellow_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=False, color=DXColor.YELLOW_TEXT), 
    fill=PatternFill(start_color=DXColor.LIGHT_YELLOW_BG, end_color=DXColor.LIGHT_YELLOW_BG),
    border=Border(left=Side(border_style='thin', color=DXColor.YELLOW_TEXT))
)

light_green_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=False, color=DXColor.GREEN_TEXT), 
    fill=PatternFill(start_color=DXColor.LIGHT_GREEN_BG, end_color=DXColor.LIGHT_GREEN_BG),
    border=Border(left=Side(border_style='thin', color=DXColor.GREEN_TEXT))
)

light_blue_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=False, color=DXColor.BLUE_TEXT), 
    fill=PatternFill(start_color=DXColor.LIGHT_BLUE_BG, end_color=DXColor.LIGHT_BLUE_BG),
    border=Border(left=Side(border_style='thin', color=DXColor.BLUE_TEXT))
)

light_gray_dx = DifferentialStyle(
    font=Font(_FONT_NAME, size=_FONT_SIZE, bold=False, color=DXColor.LIGHT_GRAY_TEXT), 
    fill=PatternFill(start_color=DXColor.LIGHT_GRAY_BG, end_color=DXColor.LIGHT_GRAY_BG),
)

