# Copyright (c) 2018-2025 Ben Fisher

from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from .dxcolor import DXColor

red_dx = DifferentialStyle(
    font=Font(name='Aptos Narrow', size=11, bold=True, color=DXColor.WHITE), 
    fill=PatternFill(start_color=DXColor.RED.replace('#',''), end_color=DXColor.RED.replace('#',''))
)

yellow_dx = DifferentialStyle(
    font=Font(name='Aptos Narrow', size=11, bold=True, color=DXColor.BLACK), 
    fill=PatternFill(start_color=DXColor.YELLOW.replace('#',''), end_color=DXColor.YELLOW.replace('#',''))
)

green_dx = DifferentialStyle(
    font=Font(name='Aptos Narrow', size=11, bold=True, color=DXColor.WHITE), 
    fill=PatternFill(start_color=DXColor.GREEN.replace('#',''), end_color=DXColor.GREEN.replace('#',''))
)

blue_dx = DifferentialStyle(
    font=Font(name='Aptos Narrow', size=11, bold=True, color=DXColor.WHITE), 
    fill=PatternFill(start_color=DXColor.BLUE.replace('#',''), end_color=DXColor.BLUE.replace('#',''))
)

gray_dx = DifferentialStyle(
    font=Font(name='Aptos Narrow', size=11, bold=True, color=DXColor.WHITE), 
    fill=PatternFill(start_color=DXColor.GRAY.replace('#',''), end_color=DXColor.GRAY.replace('#',''))
)

light_red_dx = DifferentialStyle(
    font=Font(name='Aptos Narrow', size=11, bold=True, color=DXColor.RED_TEXT.replace('#','')), 
    fill=PatternFill(start_color=DXColor.LIGHT_RED_BG.replace('#',''), end_color=DXColor.LIGHT_RED_BG.replace('#','')),
    border=Border(left=Side(border_style='thin', color=DXColor.RED_TEXT.replace('#','')))
)

light_yellow_dx = DifferentialStyle(
    font=Font(name='Aptos Narrow', size=11, bold=True, color=DXColor.YELLOW_TEXT.replace('#','')), 
    fill=PatternFill(start_color=DXColor.LIGHT_YELLOW_BG.replace('#',''), end_color=DXColor.LIGHT_YELLOW_BG.replace('#','')),
    border=Border(left=Side(border_style='thin', color=DXColor.YELLOW_TEXT.replace('#','')))
)

light_green_dx = DifferentialStyle(
    font=Font(name='Aptos Narrow', size=11, bold=False, color=DXColor.GREEN_TEXT.replace('#','')), 
    fill=PatternFill(start_color=DXColor.LIGHT_GREEN_BG.replace('#',''), end_color=DXColor.LIGHT_GREEN_BG.replace('#','')),
    border=Border(left=Side(border_style='thin', color=DXColor.GREEN_TEXT.replace('#','')))
)

light_blue_dx = DifferentialStyle(
    font=Font(name='Aptos Narrow', size=11, bold=False, color=DXColor.BLUE_TEXT.replace('#','')), 
    fill=PatternFill(start_color=DXColor.LIGHT_BLUE_BG.replace('#',''), end_color=DXColor.LIGHT_BLUE_BG.replace('#','')),
    border=Border(left=Side(border_style='thin', color=DXColor.BLUE_TEXT.replace('#','')))
)

