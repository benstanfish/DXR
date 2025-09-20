# Copyright (c) 2018-2025 Ben Fisher

from enum import IntEnum
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from .dxcolor import DXColor
from dxbuild.constants import FALLBACKS

class Widths(IntEnum):
    XX_SMALL = 5
    X_SMALL = 8
    SMALL = 10
    MEDIUM = 20
    LARGE = 50
    X_LARGE = 70

USER_NOTES_WIDTHS = [Widths.XX_SMALL, Widths.MEDIUM, Widths.MEDIUM, Widths.MEDIUM, Widths.MEDIUM, Widths.SMALL, Widths.X_SMALL]

COMMENT_COLUMN_WIDTHS = [Widths.SMALL, Widths.SMALL, Widths.MEDIUM, Widths.MEDIUM, 
                         Widths.SMALL, 
                         Widths.SMALL, Widths.MEDIUM, Widths.MEDIUM, 
                         Widths.SMALL, Widths.MEDIUM, Widths.MEDIUM, Widths.X_LARGE, 
                         Widths.SMALL, 
                         Widths.SMALL, 
                         Widths.XX_SMALL, Widths.X_SMALL, Widths.SMALL]

RESPONSE_COLUMN_WIDTHS = [Widths.SMALL, Widths.MEDIUM, 
                        #   Widths.SMALL, 
                          Widths.SMALL, Widths.X_LARGE, Widths.X_SMALL]


_FONT_NAME = FALLBACKS['font_name']
_FONT_SIZE = FALLBACKS['font_size']
_FONT_BOLD = FALLBACKS['font_bold']

project_info_key_style = {
    'font': Font(name='Aptos Narrow', size=10, bold=True),
    'fill': None,
    'border': None,
    'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False)
}

project_info_value_style = {
    'font': Font(name='Aptos Narrow', size=10, bold=False),
    'fill': None,
    'border': None,
    'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False)
}

project_title_style = {
    'font': Font(name='Aptos Narrow', size=12, bold=True),
    'fill': None,
    'border': None,
    'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False)
}

empty_status_cell_style = {
    'font': Font(name=_FONT_NAME, size=_FONT_SIZE, bold=_FONT_BOLD),
    'fill': None,
    'border': Border(left=Side(border_style='thin', color=DXColor.LIGHT_GRAY_TEXT),
                     bottom=Side(border_style='thin', color=DXColor.LIGHT_GRAY_TEXT),
                     diagonal=Side(border_style='thin', color=DXColor.LIGHT_GRAY_TEXT),
                     diagonalDown=True,
                     diagonalUp=True),
    'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False) 
}

table_header_styles = {
    'font': Font(name='Aptos Narrow', size=9, bold=True),
    'fill': None,
    'border': Border(bottom=Side(border_style='medium', color=DXColor.BLUE)),
    'alignment': Alignment(horizontal='left', vertical='bottom', wrap_text=False) 
}

table_body_styles = {
    'font': Font(name=_FONT_NAME, size=_FONT_SIZE, bold=_FONT_BOLD),
    'fill': None,
    'border': Border(bottom=Side(border_style='thin', color=DXColor.LIGHT_GRAY_TEXT)),
    'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False) 
}

table_body_wrap_styles = {
    'font': Font(name=_FONT_NAME, size=_FONT_SIZE, bold=_FONT_BOLD, ),
    'fill': None,
    'border': Border(bottom=Side(border_style='thin', color=DXColor.LIGHT_GRAY_TEXT)),
    'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True) 
}

user_notes_header_styles = {
    'font': Font(name='Aptos Narrow', size=9, bold=True, color=DXColor.WHITE),
    'fill': PatternFill(fill_type='solid', start_color=DXColor.USER_NOTES_HEADER_BG),
    'border': Border(bottom=Side(border_style='medium', color=DXColor.BLUE)),
    'alignment': Alignment(horizontal='left', vertical='bottom', wrap_text=False) 
}

user_notes_body_styles = {
    'font': Font(name=_FONT_NAME, size=_FONT_SIZE, bold=_FONT_BOLD, color=DXColor.USER_NOTES_HEADER_BG),
    'fill': PatternFill(fill_type='solid', start_color=DXColor.USER_NOTES_GRAY_BG),
    'border': Border(bottom=Side(border_style='thin', color=DXColor.GRAY)),
    'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False) 
}