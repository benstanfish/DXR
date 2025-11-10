# Copyright (c) 2018-2025 Ben Fisher

from defusedxml import ElementTree as ET
from xml.etree.ElementTree import Element
from typing import Literal
from datetime import datetime

from openpyxl.worksheet.worksheet import Worksheet

class Frameable:
    def __init__(self):
        self.frames = {}

    def shift_frames(self, col_shift: int = 0, row_shift: int = 0) -> None:
        for region in self.frames:
            if self.frames[region] is not None:
                self.frames[region].shift(col_shift=col_shift, row_shift=row_shift)

    def expand_frame(self, frame_name:str, right:int=0, down:int=0, left:int=0, up:int=0):
        if self.frames[frame_name] is not None:
            self.frames[frame_name].expand(right=right, down=down, left=left, up=up)  

    def get_anchor_cell(self, worksheet:Worksheet, frame_name:str=''):
        if self.frames[frame_name] is not None:
            if frame_name:
                return worksheet.cell(row=self.frames[frame_name].min_row, column=self.frames[frame_name].min_col).coordinate
            else:
                if 'extents' in self.frames.keys():
                    return worksheet.cell(row=self.frames['extents'].min_row, column=self.frames['extents'].min_col).coordinate
                return worksheet.cell(row=self.frames[0].min_row, column=self.frames[0].min_col).coordinate
        