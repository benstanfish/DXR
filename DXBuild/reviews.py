# Copyright (c) 2018-2025 Ben Fisher

from os.path import getctime
from datetime import datetime
from typing import List, Dict, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.cell_range import CellRange

from .constants import (COMMENT_COLUMNS, 
                        RESPONSE_COLUMNS, 
                        USER_NOTES_COLUMNS,
                        _PROJECT_INFO_INDEX,
                        _COMMENTS_INDEX,
                        _RESPONSE_EXPANSION_TYPES)
from .parseable import Parseable
from .remarks import Comment

class Frameable:
    def __init__(self):
        self.frames = {}

    def shift_frames(self, col_shift: int = 0, row_shift: int = 0) -> None:
        for region in self.frames:
            if self.frames[region] is not None:
                self.frames[region].shift(col_shift=col_shift, row_shift=row_shift)

    def expand_frame(self, frame_name:str, right:int=0, down:int=0, left:int=0, up:int=0):
        if self.frames[frame_name] is not None:
            self.frames[frame_name].expand(right=right, down=down, left=left, up=up)  

    def get_anchor_cell(self, worksheet:Worksheet, frame_name:str=''):
        if self.frames[frame_name] is not None:
            if frame_name:
                return worksheet.cell(row=self.frames[frame_name].min_row, column=self.frames[frame_name].min_col).coordinate
            else:
                if 'extents' in self.frames.keys():
                    return worksheet.cell(row=self.frames['extents'].min_row, column=self.frames['extents'].min_col).coordinate
                return worksheet.cell(row=self.frames[0].min_row, column=self.frames[0].min_col).coordinate
        

class ProjectInfo(Frameable, Parseable):

    def __init__(self, 
                 project_id=None,
                 control_number=None,
                 project_name=None,
                 review_id=None,
                 review_name=None,
                 xml_date=None,
                 run_date=None):
        super().__init__()
        self.project_id = project_id
        self.control_number = control_number
        self.project_name = project_name
        self.review_id = review_id
        self.review_name = review_name
        self.xml_date = xml_date
        self.run_date = run_date
        self.create_initial_frames()
        
    @classmethod
    def from_element(cls, element, file_path=None):
        project_id = cls.parse_single_tag('ProjectID', element)
        control_number = cls.parse_single_tag('ProjectControlNbr', element)
        project_name = cls.parse_single_tag('ProjectName', element)
        review_id = cls.parse_single_tag('ReviewID', element)
        review_name = cls.parse_single_tag('ReviewName', element)
        if file_path is not None:
            file_date = getctime(file_path)
            xml_date = datetime.fromtimestamp(file_date).strftime('%Y-%m-%d %H:%M:%S')
        else:
            xml_date = None
        run_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return ProjectInfo(project_id=project_id,
                           control_number=control_number,
                           project_name=project_name,
                           review_id=review_id,
                           review_name=review_name,
                           xml_date=xml_date,
                           run_date=run_date)
   
    @property
    def all_properties_to_dict(self) -> Dict:
        return {
            'Project ID': self.project_id,
            'Control Number': self.control_number,
            'Project Name': self.project_name,
            'Review Name': self.review_name,
            'Review ID': self.review_id,
            'XML Date': self.date_to_excel(self.xml_date),
            'Run Date': self.date_to_excel(self.run_date),
            'Notes': ''
        }
    
    def to_list(self) -> List:
        """Returns a 2D list intended to be exported to Excel."""
        info = []
        for key in self.all_properties_to_dict.keys():
            info.append([key, self.all_properties_to_dict[key]])
        return info

    @property
    def count(self) -> int:
        return len(self.to_list())

    @property
    def size(self) -> Tuple[int, int]:
        return (self.count, 2)
    
    def create_initial_frames(self) -> None:
        """Set the initial cell range frames that define the Project Info region.

        These 'frames' are Openpyxl CellRanges that define regions for the Project Info data
        in a Worksheet. Initially, the Project Info frame is created starting at 'A1' however,
        this will need to be shifted, once the User Notes object is created, given that is located
        to the left of the Project Info and Comments table.
        """
        self.frames['extents'] = CellRange(min_col=1, max_col=2, min_row=1, max_row=self.count)
        self.frames['keys'] = CellRange(min_col=1, max_col=1, min_row=1, max_row=self.count)
        self.frames['values'] = CellRange(min_col=2, max_col=2, min_row=1, max_row=self.count)
        self.frames['project_title'] = CellRange(min_col=2, max_col=2, min_row=3, max_row=3)



class ReviewComments(Frameable, Parseable):

    def __init__(self,
                 comments=[]):
        super().__init__()
        self.comments = comments
        self.create_initial_frames()

    @classmethod
    def from_tree(cls, element):
        comments = []
        for comment in element.findall('comment'):
            comments.append(Comment.from_element(comment))
        return ReviewComments(comments=comments)
    
    @property
    def count(self) -> int:
        return len(self.comments)

    @property
    def max_evaluations(self) -> int:
        temp_count = 0
        for comment in self.comments:
            if comment.evaluations_count > temp_count:
                temp_count = comment.evaluations_count 
        return temp_count

    @property
    def max_backchecks(self) -> int:
        temp_count = 0
        for comment in self.comments:
            if comment.backchecks_count > temp_count:
                temp_count = comment.backchecks_count 
        return temp_count    

    @property
    def max_responses(self) -> Tuple[int, int]:
        return (self.max_evaluations, 
                self.max_backchecks)

    @property
    def evaluations_count(self) -> int:
        temp_count = 0
        for comment in self.comments:
            temp_count += comment.evaluations_count
        return temp_count

    @property
    def backchecks_count(self) -> int:
        temp_count = 0
        for comment in self.comments:
            temp_count += comment.backchecks_count
        return temp_count
    
    @property
    def responses_count(self) -> int:
        return self.evaluations_count + self.backchecks_count

    @property
    def comments_to_list(self, attrs: Dict=COMMENT_COLUMNS) -> List:
        my_data = []
        for comment in self.comments:
            my_data.append(comment.to_list(attrs))
        return my_data

    @property
    def column_names(self, attrs: Dict=COMMENT_COLUMNS) -> List:
        return [key for key in attrs.keys()]

    @property
    def comment_columns_count(self, attrs: Dict=COMMENT_COLUMNS) -> int:
        return len([key for key in attrs.keys()])   

    @property
    def response_columns_count(self, attrs: Dict=RESPONSE_COLUMNS) -> int:
        return len([key for key in attrs.keys()] * (self.max_evaluations + self.max_backchecks))
    
    @property
    def all_column_count(self):
        return self.comment_columns_count + self.response_columns_count

    def _get_all_headers(self,
                        comment_attrs: Dict=COMMENT_COLUMNS,
                        response_attrs: Dict=RESPONSE_COLUMNS,
                        expansion_type: _RESPONSE_EXPANSION_TYPES ='chronological') -> List:
        """Returns a list of all the column names (for use in Excel) and the column count."""
        header_names = [key for key in comment_attrs.keys()]
        if self.responses_count > 0:
            max_evals, max_bcs = self.max_responses
            if expansion_type.lower() != 'chronological':
                for i in range(max_evals):
                    for key in response_attrs.keys():
                        header_names.append(f'Eval {i + 1} {key}')
                for j in range(max_bcs):
                    for key in response_attrs.keys():
                        header_names.append(f'BCheck {j + 1} {key}')
            else:
                for k in range(max_evals + max_bcs):
                    for key in response_attrs.keys():
                        header_names.append(f'Resp {k + 1} {key}')
        return header_names

    def to_list(self, 
                comment_attrs: Dict=COMMENT_COLUMNS,
                response_attrs: Dict=RESPONSE_COLUMNS,
                expansion_type: _RESPONSE_EXPANSION_TYPES='chronological') -> List:
        """Returns the full List of comments and corresponding responses and the number of rows."""

        full_list = [self._get_all_headers(comment_attrs=comment_attrs, 
                                          response_attrs=response_attrs, 
                                          expansion_type=expansion_type)]
        max_eval_count, max_bc_count = self.max_responses
        for comment in self.comments:
            temp = []
            temp += comment.to_list(comment_attrs)
            if self.responses_count > 0:
                if expansion_type == 'chronological':
                    resp_list = comment.list_responses_chronological
                    resp_count = comment.total_response_count
                    diff_eval = (max_eval_count + max_bc_count) - resp_count
                    for resp in resp_list:
                        temp += resp.to_list(response_attrs)
                    for i in range(diff_eval):
                        temp += [''] * len(response_attrs)
                    full_list.append(temp)
                else:
                    for evaluation in comment.evaluations:
                        temp += evaluation.to_list(response_attrs)
                    diff_eval = max_eval_count - comment.evaluations_count
                    for _ in range(diff_eval):
                        temp += [''] * len(response_attrs)
                    for backcheck in comment.backchecks:
                        temp += backcheck.to_list(response_attrs)
                    diff_bc = max_bc_count - comment.backchecks_count
                    for _ in range(diff_bc):
                        temp += [''] * len(response_attrs)
                    full_list.append(temp)
            else:
                full_list.append(temp)
        return full_list

    @property
    def size(self) -> Tuple[int, int]:
        return (self.count, self.all_column_count)

    def create_initial_frames(self) -> None:
        """Set the initial cell range frames that define the Review Comments region.

        These 'frames' are Openpyxl CellRanges that define regions for the Review Comments data
        in a Worksheet. Initially, the Project Info frame is created starting at 'A1' however,
        this will need to be shifted, once the Project Info and User Notes objects are created.
        """
        self.frames['extents'] = CellRange(min_col=1, max_col=self.all_column_count, min_row=1, max_row=self.count + 1)
        self.frames['header'] = CellRange(min_col=1, max_col=self.all_column_count, min_row=1, max_row=1)
        self.frames['body'] = CellRange(min_col=1, max_col=self.all_column_count, min_row=2, max_row=self.count + 1)
        self.frames['comments_header'] = CellRange(min_col=1, max_col=self.comment_columns_count, min_row=1, max_row=1)
        self.frames['comments_body'] = CellRange(min_col=1, max_col=self.comment_columns_count, min_row=2, max_row=self.count + 1)
        if self.responses_count > 0:
            self.frames['response_header'] = CellRange(min_col=1 + self.comment_columns_count, max_col=self.comment_columns_count + self.response_columns_count, min_row=1, max_row=1)
            self.frames['response_body'] = CellRange(min_col=1 + self.comment_columns_count, max_col=self.comment_columns_count + self.response_columns_count, min_row=2, max_row=self.count + 1)


class UserNotes(Frameable):
    def __init__(self):
        super().__init__()
        self.headers = [header for header in USER_NOTES_COLUMNS]
        self.create_initial_frames()

    def to_list(self) -> List[str]:
        return self.headers
    
    @property
    def size(self) -> Tuple[int, int]:
        return (1, self.count)

    @property
    def count(self) -> int:
        return len(self.headers)
    
    def create_initial_frames(self) -> None:
        """Set the initial cell range frames that define the User Notes region.

        These 'frames' are Openpyxl CellRanges that define regions for the User Note data
        in a Worksheet. Initially, the Project Info frame is created starting at 'A1' however,
        this will need to be shifted, once the Project Info object is created. This will determine
        how far down to shift the region. Also, the Review Comments region will specify how
        many rows to increase the User Notes body region.
        """
        self.frames['extents'] = CellRange(min_col=1, max_col=self.count, min_row=1, max_row=2)
        self.frames['header'] = CellRange(min_col=1, max_col=self.count, min_row=1, max_row=1)
        self.frames['body'] = CellRange(min_col=1, max_col=self.count, min_row=2, max_row=2)
        self.frames['id_column'] = CellRange(min_col=1, max_col=1, min_row=2, max_row=2)
    
    def shift_frames(self, col_shift: int = 0, row_shift: int = 0) -> None:
        for region in self.frames:
            self.frames[region].shift(col_shift=col_shift, row_shift=row_shift)  

    def expand_frame(self, frame_name:str, right:int=0, down:int=0, left:int=0, up:int=0):
        self.frames[frame_name].expand(right=right, down=down, left=left, up=up)        

    def autonumber_id_column(self, worksheet:Worksheet) -> None:
        i = 1
        for row in worksheet[self.frames['id_column'].coord]:
            for cell in row:
                cell.value = i
                i += 1

class Review(Frameable, Parseable):
    """Returns a Review object containing project info and review comments objects."""

    def __init__(self,
                 project_info: ProjectInfo,
                 review_comments: ReviewComments,
                 root=None,
                 file_path=None):
        super().__init__()
        self.project_info = project_info
        self.review_comments = review_comments
        self.root = root
        self.file_path = file_path
        self.user_notes = UserNotes()
        self.table_column_list = []
        self.setup_frames()

    @classmethod
    def from_file(cls, path):
        root = cls.get_root(path) if not None else None
        if root:
            project_info = ProjectInfo.from_element(root[_PROJECT_INFO_INDEX], file_path=path) if not None else None
            review_comments = ReviewComments.from_tree(root[_COMMENTS_INDEX]) if not None else None
        return Review(project_info=project_info,
                      review_comments=review_comments,
                      root=root,
                      file_path=path)

    def setup_frames(self):
        project_info_height = self.project_info.count
        project_info_offset = 2
        comment_rows = self.review_comments.count
        table_header_row = project_info_height + project_info_offset
        table_header_column = self.user_notes.count
        self.project_info.shift_frames(col_shift=table_header_column)
        self.user_notes.shift_frames(row_shift=table_header_row)
        self.user_notes.expand_frame('extents', down=comment_rows-1)
        self.user_notes.expand_frame('body', down=comment_rows-1)
        self.user_notes.expand_frame('id_column', down=comment_rows-1)
        self.review_comments.shift_frames(col_shift=table_header_column, row_shift=table_header_row)
        self.frames['extents'] = self.user_notes.frames['extents'].union(self.review_comments.frames['extents'])
        self.frames['header'] = self.user_notes.frames['header'].union(self.review_comments.frames['header'])
        self.frames['body'] = self.user_notes.frames['body'].union(self.review_comments.frames['body'])

    def build_table_column_list(self, worksheet:Worksheet) -> List:
        temp = []
        for row in worksheet[self.frames['header'].coord]:
            for cell in row:
                if cell.value not in temp:
                    temp.append(cell.value)
        self.table_column_list = temp

    @property
    def is_valid(self):
        if self.root.tag.lower() == 'projnet':
            return True
        return False