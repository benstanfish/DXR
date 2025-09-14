# Copyright (c) 2018-2025 Ben Fisher

from datetime import datetime
from typing import List, Tuple, Dict, Literal

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import CellRange




#TODO: Need to split the DrChecks Review formatting tools from the basic range operations tools (different modules)


def timestamp(
        format_string: str=r'%Y%m%d_%H%M%S'
    ) -> str:
    """Returns a formatted timestamp of the current time.

    :param format_string: format string provided in accordance with https://docs.python.org/3/library/datetime.html#format-codes, defaults to r'%Y%m%d_%H%M%S'
    :type format_string: str, optional
    :return: current time as a timestamp string.
    :rtype: str
    """
    # """Returns a timestamp, default format: YYYYMMDD_HHMMSS"""
    return datetime.now().strftime(format_string)


def list_dimensions(
        input_list: List
    ) -> Tuple[int, int]:
    """Returns a Tuple of (Row, Column) count, assuming 1D or 2D Lists."""
    if isinstance(input_list[0], list):
        rows = len(input_list)
        cols = len(input_list[0]) if isinstance(input_list[0], list) else 1
    else:
        rows = 1
        cols = len(input_list)
    return (rows, cols)
    

def copy_to_range(
        data_list: List, 
        worksheet: Worksheet | None, 
        anchor_cell: str='A1'
    ) -> None:
    """Copy a 1D or 2D list to worksheet via Openpyxl."""
    if worksheet is not None:
        anchor_row, anchor_column = coordinate_to_tuple(anchor_cell)
        rows, columns = list_dimensions(data_list)
        if rows == 1:
            for j in range(columns):
                worksheet.cell(row=anchor_row, column=anchor_column + j).value = data_list[j]
        else:
            for i in range(rows):
                for j in range(columns):
                    worksheet.cell(row=anchor_row + i, column=anchor_column + j).value = data_list[i][j]


def range_string_from_bounds(
        min_col: int, 
        max_col: int,
        min_row: int, 
        max_row: int, 
        min_col_abs: bool=False,
        max_col_abs: bool=False,
        min_row_abs: bool=False,
        max_row_abs: bool=False
    ) -> str:
    """Return range string from row, column bounds. Allows for partwise absolute referencing."""
    start_col_letter = get_column_letter(min_col)
    end_col_letter = get_column_letter(max_col)
    
    return f'{"$" if min_col_abs else ""}{start_col_letter}' \
           f'{"$" if min_row_abs else ""}{min_row}' \
           f':{"$" if max_col_abs else ""}{end_col_letter}' \
           f'{"$" if max_row_abs else ""}{max_row}'


def bounds_from_range_string(
        range: str | CellRange
    ) -> tuple:
    """Returns the tuple of row, column min, max bounds from a CellRange or range-like string."""
    if isinstance(range, CellRange):
        return (range.min_col, range.max_col, range.min_row, range.max_row)
    elif isinstance(range, str):
        # Test to see if the passed string is 'range-like'
        import re
        range_like = r'\b[a-zA-Z]{1,3}\d{1,7}\b'
        matches = re.findall(pattern=range_like, string=range)
        if matches:
            # This is the case where the string appears to be 'range-like'
            # convert to a CellRange and extract the bounds
            temp = CellRange(range)
            return (temp.min_col, temp.max_col, temp.min_row, temp.max_row)
    # If code continues to this point, the input was invalid and an empty tuple is returned.
    return ()


def start_end_cells_from_range(
        range_string: str | CellRange
    ) -> List:
    """Returns a list of the first and last cells in a range as strings."""
    if isinstance(range_string, CellRange):
        return [f'{get_column_letter(range_string.min_col)}{range_string.min_row}',
                f'{get_column_letter(range_string.max_col)}{range_string.max_row}']
    elif isinstance(range_string, str):
        # Test to see if the passed string is 'range-like'
        import re
        range_pattern = r'\b([a-zA-Z]{1,3}\d{1,7})(:[a-zA-Z]{1,3}\d{1,7}){0,1}\b'
        matches = re.findall(pattern=range_pattern, string=range_string)
        if matches:
            # This is the case where the string appears to be 'range-like'
            # convert to a CellRange and extract the bounds
            temp = CellRange(range_string)
            return [f'{get_column_letter(temp.min_col)}{temp.min_row}',
                    f'{get_column_letter(temp.max_col)}{temp.max_row}']
    # If code continues to this point, the input was invalid and an empty tuple is returned.
    return []


abs_rel_type = Literal['column', 'row', 'both', 'none']
def abs_rel_address(
        range_string: str, 
        type:abs_rel_type='none'
    ) -> str:
    """Returns range-like string with absolute or relative """
    temp = range_string.replace('$','')
    import re
    range_like = r'\b([a-zA-Z]{1,3})(\d{1,7})\b'
    match = re.match(pattern=range_like, string=temp)
    if match:
        whole_match = match.group(0)    # The whole match (not used?)
        letters = match.group(1)        # Group 1 of the match
        numbers = match.group(2)        # Group 2 of the match
        if type == 'both':
            return '$' + letters + '$' + numbers
        elif type == 'column':
            return '$' + letters + numbers
        elif type == 'row':
            return letters + '$' + numbers
        else:
            return whole_match
    return ''
    

def autoincrement_name(
        base_name: str, 
        search_list: List
    ) -> str:
    """_summary_

    :param base_name: _description_
    :type base_name: str
    :param search_list: _description_
    :type search_list: List
    :return: _description_
    :rtype: str
    """
    # """Returns the base_name with next largest integer suffixed if name already exists in search_list."""
    base_name_length = len(base_name)
    temp = []
    for thing in search_list:
        if thing[:base_name_length] == base_name:
            temp.append(thing)
    if len(temp):
        # This block executes if there are potential collisions with the base name
        max_number = 0
        for thing in temp:
            if len(thing) > base_name_length:
                curr_num = int(thing[base_name_length:])
                if curr_num > max_number:
                    max_number = curr_num
        return base_name + str(max_number + 1)
    # If no collisions, the base name is returned.
    return base_name

def add_data_validation_to_column(
    options_string: str,
    cell_range_list: List[str], 
    worksheet: Worksheet
    ) -> None:
    """_summary_

    :param options_string: _description_
    :type options_string: str
    :param worksheet: _description_
    :type worksheet: Worksheet
    :param allow_blank: _description_, defaults to True
    :type allow_blank: bool, optional
    """
    dv = DataValidation(type='list', 
                        formula1=f'"{options_string}"', 
                        allow_blank=True)
    worksheet.add_data_validation(dv)
    for cell_range in cell_range_list:
        dv.add(CellRange(cell_range))


def conditionally_format_range(
        check_range_string: str, 
        check_for_string: str, 
        ws: Worksheet, 
        dxf: DifferentialStyle,
        apply_to_range_string: str='',
        stop_if_true:bool = False
    ) -> None:
    start_cell, end_cell = start_end_cells_from_range(check_range_string)
    if start_cell:
        formula_string = [f'=LOWER({abs_rel_address(range_string=start_cell, 
                          type='column')})="{check_for_string}"']
        cf_rule = Rule(type='expression', 
                        dxf=dxf, 
                        formula=formula_string, 
                        stopIfTrue=stop_if_true)
        if apply_to_range_string:
            ws.conditional_formatting.add(range_string=apply_to_range_string, cfRule=cf_rule)
        else:
            ws.conditional_formatting.add(range_string=check_range_string, cfRule=cf_rule)


def get_table_info(worksheet: Worksheet) -> Dict:
    """Returns a dictionary containing the header row number and first and last rows of the databody range of the first listobject in the supplied Worksheet.

    :param worksheet: _description_
    :type worksheet: Worksheet
    :return: _description_
    :rtype: List[int]
    """
    table_data = worksheet.tables.items()
    min_col, max_col, min_row, max_row = bounds_from_range_string(table_data[0][1])
    header_cell_range = CellRange(min_col=min_col, max_col=max_col, min_row=min_row, max_row=min_row)
    header_names = [header.value for header in worksheet[header_cell_range.coord][0]]
    return {'header_row': min_row, 'first_row': min_row + 1, 'last_row': max_row, 'headers': header_names}


def list_column_range(column_index_or_letter: str | int, table_info_dict: dict) -> str:
    """Returns the range address for the databodyrange of a given column.

    :param column_index_or_letter: _description_
    :type column_index_or_letter: str | int
    :param table_row_info: _description_
    :type table_row_info: dict
    :return: _description_
    :rtype: str
    """
    header_row, first_row, last_row = table_info_dict['header_row'], table_info_dict['first_row'], table_info_dict['last_row']
    column_letter = ''
    if isinstance(column_index_or_letter, int):
        column_letter = get_column_letter(column_index_or_letter)
    if isinstance(column_index_or_letter, str):
        column_letter = column_index_or_letter.upper()
    return f'{column_letter}{first_row}:{column_letter}{last_row}'


def get_columns_by_name(search_text: str, table_info_dict: Dict) -> List[str]:
    """Returns a list of the worksheet column letters for each column in a table that includes the search text.

    :param search_text: _description_
    :type search_text: str
    :param table_info_dict: _description_
    :type table_info_dict: Dict
    :return: _description_
    :rtype: List[str]
    """
    list_columns =table_info_dict['headers']
    temp = []
    for header_no, header in enumerate(list_columns):
        if search_text in header.lower():
            temp.append(get_column_letter(header_no + 1))
    return temp


def build_column_vectors(column_list: list[str], table_info_dict: Dict) -> List[str]:
    """Returns a list of column ranges for a table rows (ignoring the header row).

    :param column_list: _description_
    :type column_list: list[str]
    :param table_info_dict: _description_
    :type table_info_dict: Dict
    :return: _description_
    :rtype: List[str]
    """
    return [list_column_range(column, table_info_dict) for column in column_list]