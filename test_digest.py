from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import DEFAULT_FONT

from dxbuild.reviews import Review
import dxbuild.buildtools as buildtools
from dxbuild.constants import RESPONSE_COLUMNS, USER_NOTES_COLUMNS
from dxcore.conditionalformats import *
from dxcore.cellformats import *
from dxbuild.constants import FALLBACKS

# Debug information
_WRITE_FILE = True
xml_path = './dev/test/data.xml'

# Parse XML and create data objects
review = Review.from_file(xml_path)
project_info = review.project_info
all_comments = review.review_comments
user_notes = review.user_notes

# Create workbook object with initial settings
DEFAULT_FONT.__init__(name=FALLBACKS['font_name'], size=FALLBACKS['font_size'])
wb = Workbook()
ws = wb.active

# Get upper left cell of each major region
user_notes_anchor_cell = review.user_notes.get_anchor_cell(ws, 'extents')
project_info_anchor_cell = review.project_info.get_anchor_cell(ws, 'extents')
review_comments_anchor = review.review_comments.get_anchor_cell(ws, 'extents')

# Dump Data to Excel
buildtools.copy_to_range(user_notes.to_list(), worksheet=ws, anchor_cell=user_notes_anchor_cell)
buildtools.copy_to_range(project_info.to_list(), worksheet=ws, anchor_cell=project_info_anchor_cell)
buildtools.copy_to_range(all_comments.to_list(), worksheet=ws, anchor_cell=review_comments_anchor)

user_notes.autonumber_id_column(ws)

TABLE_RANGE = review.frames['extents'].coord
TABLE_HEADER = review.frames['header'].coord
TABLE_BODY = review.frames['body'].coord
table = Table(displayName='Comments', ref=TABLE_RANGE)

if ws is not None:
    ws.add_table(table)
    ws.sheet_view.showGridLines = False
    table_info = buildtools.get_table_info(ws)
    
    review.build_table_column_list(ws)
    table_column_list = review.table_column_list

    top_left_alignment = Alignment(horizontal='left', vertical='top')
    used_range = ws.calculate_dimension()
    for row in ws[used_range]:
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'm/d/yy'
            cell.alignment = top_left_alignment

    buildtools.apply_styles_to_region(table_header_styles, TABLE_HEADER, ws)
    buildtools.apply_styles_to_region(table_body_styles, TABLE_BODY, ws)
    buildtools.apply_styles_to_region(user_notes_header_styles, review.user_notes.frames['header'].coord, ws)
    buildtools.apply_styles_to_region(user_notes_body_styles, review.user_notes.frames['body'].coord, ws)

    open_closed_options = 'Open, Closed'
    status_options = 'Concur, For Information Only, Non-Concur, Check and Resolve'
    
    status_column_letters = buildtools.get_columns_by_name('status', table_info)
    status_columns = buildtools.build_column_vectors(status_column_letters, table_info) # Treat the first occurance separately
    buildtools.add_data_validation_to_column(status_options, status_columns[1:], ws)
    
    first_status_column = buildtools.add_data_validation_to_column(open_closed_options, status_columns[:1], ws)
    buildtools.conditionally_format_range(status_columns[0], 
                                          'closed', 
                                          ws, 
                                          light_gray_dx, 
                                          review.review_comments.frames['body'].coord, 
                                          stop_if_true=True)
        
    for column in status_columns[1:]:
        buildtools.conditionally_format_range(column, 'check and resolve', ws, light_red_dx)
        buildtools.conditionally_format_range(column, 'non-concur', ws, light_yellow_dx)
        buildtools.conditionally_format_range(column, 'for information only', ws, light_green_dx)
        buildtools.conditionally_format_range(column, 'concur', ws, light_blue_dx)
        buildtools.apply_styles_to_region_if_empty(empty_status_cell_style, column, ws)
    
    highest_reponse_letters = buildtools.get_columns_by_name('highest resp', table_info)
    highest_response_columns = buildtools.build_column_vectors(highest_reponse_letters, table_info)[0]
    buildtools.add_data_validation_to_column(status_options, [highest_response_columns], ws)
    buildtools.conditionally_format_range(highest_response_columns, 'check and resolve', ws, red_dx)
    buildtools.conditionally_format_range(highest_response_columns, 'non-concur', ws, yellow_dx)
    buildtools.conditionally_format_range(highest_response_columns, 'for information only', ws, green_dx)
    buildtools.conditionally_format_range(highest_response_columns, 'concur', ws, blue_dx)
    buildtools.apply_styles_to_region_if_empty(empty_status_cell_style, highest_response_columns, ws)

    critical_options = 'Yes, No'
    critical_column_letters = buildtools.get_columns_by_name('critical', table_info)
    critical_column = buildtools.build_column_vectors(critical_column_letters, table_info)
    buildtools.add_data_validation_to_column(critical_options, critical_column, ws)
    buildtools.conditionally_format_range(critical_column[0], 'yes', ws, red_dx)    
    
    class_options = 'CUI, Unclassified, Public'
    class_column_letters = buildtools.get_columns_by_name('class', table_info)
    class_column = buildtools.build_column_vectors(class_column_letters, table_info)
    buildtools.add_data_validation_to_column(class_options, class_column, ws)
    buildtools.conditionally_format_range(class_column[0], 'cui', ws, red_dx)   
    buildtools.conditionally_format_range(class_column[0], 'unclassified', ws, yellow_dx)


    
    for i, col_width in enumerate(USER_NOTES_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = col_width

    for i, col_width in enumerate(COMMENT_COLUMN_WIDTHS):
        ws.column_dimensions[get_column_letter(i + review.user_notes.count + 1)].width = col_width

    #TODO: Work through the logic of setting widths for the Response columns
    response_header_cell_range = review.review_comments.frames['response_header']
    for i in range(response_header_cell_range.min_col, response_header_cell_range.max_col + 1):
        i_modulo = (i - response_header_cell_range.min_col) % len(RESPONSE_COLUMNS)
        ws.column_dimensions[f'{get_column_letter(i)}'].width = RESPONSE_COLUMN_WIDTHS[i_modulo]

    wrappables = USER_NOTES_COLUMNS + ['Comment']
    #FIXME add 'Text' to the list above to wrap the Text column in the responses. Right now
    #it's not wrapping because Openpyxl column_dimensions.group changes reverts the column widths
    #and I need to figure out how to prevent that from happening.
    for wrappable in wrappables:
        for item in review.table_column_list:
            if wrappable in item:
                range_string = f'{get_column_letter(review.table_column_list.index(item) + 1)}{review.frames['body'].min_row}:' \
                               f'{get_column_letter(review.table_column_list.index(item) + 1)}{review.frames['body'].max_row}'
                buildtools.apply_styles_to_region(table_body_wrap_styles, range_string, ws)

    # Group columns
    collapse_left = '▷'
    collapse_right = '◁ Expand'
    collapse_regions = [['Notes', 'State'],
                        ['Source', 'Section']]

    response_collapse_region = []   # This is used in conjunction wiht the reponse_header_cell_range from above
    for col in range(response_header_cell_range.min_col, response_header_cell_range.max_col + 1):
        current_modulo = (col - response_header_cell_range.min_col) % len(RESPONSE_COLUMNS)
        if current_modulo == 1 or current_modulo == len(RESPONSE_COLUMNS) - 1:
            response_collapse_region.append(table_column_list[col - 1])
        if current_modulo == len(RESPONSE_COLUMNS) - 1:
            collapse_regions.append(response_collapse_region)
            response_collapse_region = []
        
    for region in collapse_regions:
        start_col = table_column_list.index(region[0]) + 1
        stop_col = table_column_list.index(region[1]) + 1

        label_left_cell = ws[f'{get_column_letter(start_col - 1)}{review.frames['header'].min_row - 1}']
        label_left_cell.value = collapse_left
        label_left_cell.alignment = Alignment(horizontal='right', vertical='top')
        label_right_cell = ws[f'{get_column_letter(stop_col + 1)}{review.frames['header'].min_row - 1}']
        label_right_cell.value = collapse_right

        ws.column_dimensions.group(get_column_letter(start_col), 
                                   get_column_letter(stop_col), 
                                   hidden=True)

    # for col_dim in ws.column_dimensions.values():
    #     if col_dim.outlineLevel > 0:
    #         col_dim.hidden = True
    #         col_dim.collapsed = True


if _WRITE_FILE:
    save_name = f'./dev/test/out/test_{buildtools.timestamp()}.xlsx'
    wb.save(save_name)
    print(f'{save_name} has been saved.')
wb.close()