from openpyxl import Workbook
from openpyxl.styles import PatternFill, Fill, Border, Side, Alignment, NamedStyle, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

file_path = './output.xlsx'

wb = Workbook()
ws = wb.create_sheet('create_sheet', 0)
ws.title = 'sheet_title'

tree_data = [["Type", "Leaf Color", "Height"], 
            ["Maple", "Red", 549], 
            ["Oak", "Green", 783], 
            ["Pine", "Green", 1204]]

for row in tree_data:
    ws.append(row)

# left_top_alignment = Alignment(horizontal='left',
#                                vertical='top',
#                                wrapText=True)

# font = Font(name='Arial',
#             size=10,
#             bold=True,
#             italic=False,
#             vertAlign=None,
#             underline='none',
#             strike=False,
#             color='00CC99FF')

# ws['A1'].font = font
# header_range = ws['C1']
# header_range.font = Font(name='Aptos', bold=True, color='FF000000')


# named_style = NamedStyle(name='my_style',
#                         font=Font(name='Aptos', size=20, bold=True, color='00FF0000'))

# wb.add_named_style(named_style)
# ws['b3'].style = 'my_style'

my_table = Table(displayName='my_table', ref='A1:C4')
# has_no_style = True
# if has_no_style:
#     style = TableStyleInfo(name='TableStyleMedium9', 
#                            showFirstColumn=False,
#                            showLastColumn=False,
#                            showRowStripes=False,
#                            showColumnStripes=False)
#     my_table.tableStyleInfo = style

# There's basically no reason to apply a TableStyleInfo because that
# simply adds a default table style.


ws.add_table(my_table)





wb.save(file_path)
wb.close()