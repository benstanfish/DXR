from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, DEFAULT_FONT
import openpyxl.utils

file_path = './output.xlsx'

DEFAULT_FONT.__init__(name='Aptos', size=10.5)


wb = Workbook()

# wb._fonts[0].name = 'Arial'

ws = wb.active
ws.title = 'sheet_title'

tree_data = [["Type", "Leaf Color", "Height"],
            ["Maple", "Red", 549],
            ["Oak", "Green", 783],
            ["Pine", "Green", 1204]]

for row in tree_data:
    ws.append(row)

# normal_font = Font(name='Aptos',
#                    size=11)

# header_font = Font(name='Aptos',
#                    size=11,
#                    bold=True)

# regular_align = Alignment(horizontal='left',
#                           vertical='top')

# all_cells = ws['A1:C4']
# for row in all_cells:
#     for cell in row:
#         cell.font = normal_font
#         cell.alignment = regular_align

# header_range = ws['A1:C1']
# for row in header_range:
#     for cell in row:
#         cell.font = header_font
#         cell.alignment = regular_align




wb.save(file_path)
wb.close()